"""
M2R CRM UI Components
=====================
Tkinter UI components for the M2R CRM application.
"""

import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, date
from typing import Callable, Optional, Dict, Any, List

import glob as _glob
import os
import threading

from m2r_crm_paths import APP_DIR

try:
    import send_newsletter as _newsletter_mod
    _NEWSLETTER_AVAILABLE = True
except Exception:
    _newsletter_mod = None
    _NEWSLETTER_AVAILABLE = False

import json as _json

def _load_config() -> dict:
    path = os.path.join(str(APP_DIR), "m2r_crm_config.json")
    try:
        with open(path, "r", encoding="utf-8") as _f:
            return _json.load(_f)
    except Exception:
        return {}

def _save_config(data: dict) -> None:
    path = os.path.join(str(APP_DIR), "m2r_crm_config.json")
    try:
        with open(path, "w", encoding="utf-8") as _f:
            _json.dump(data, _f, indent=2)
    except Exception:
        pass

try:
    from tkcalendar import Calendar as _TkCalendar

    class DateEntry(ttk.Frame):
        """
        Custom date entry with a popup calendar.
        Replaces tkcalendar.DateEntry to fix two bugs:
          - Calendar popup cut off at screen edges
          - Clicking month navigation arrows closing the picker
        """

        def __init__(self, parent, width=14, date_pattern='mm/dd/yyyy', **kwargs):
            super().__init__(parent)
            self._popup = None
            self._var = tk.StringVar()
            self._entry = ttk.Entry(self, textvariable=self._var, width=width)
            self._entry.pack(side="left")
            ttk.Button(self, text="▼", width=2, command=self._toggle).pack(side="left")

        # ── Public API matching tkcalendar.DateEntry ──────────────────────────

        def get(self):
            return self._var.get()

        def get_date(self):
            val = self._var.get().strip()
            for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(val, fmt).date()
                except ValueError:
                    pass
            return None

        def set_date(self, d):
            if hasattr(d, 'strftime'):
                self._var.set(d.strftime('%m/%d/%Y'))

        def delete(self, start, end):
            self._var.set('')

        def insert(self, idx, value):
            self._var.set(value)

        # ── Private ───────────────────────────────────────────────────────────

        def _toggle(self):
            if self._popup and self._popup.winfo_exists():
                self._close()
            else:
                self._open()

        def _close(self):
            if self._popup and self._popup.winfo_exists():
                self._popup.destroy()
            self._popup = None

        def _open(self):
            self._popup = tk.Toplevel(self)
            self._popup.wm_overrideredirect(True)
            self._popup.lift()

            cur = self.get_date()
            if cur:
                kw = dict(year=cur.year, month=cur.month, day=cur.day)
            else:
                t = date.today()
                kw = dict(year=t.year, month=t.month)

            cal = _TkCalendar(self._popup, selectmode='day', **kw)
            cal.pack(padx=4, pady=4)
            if cur:
                cal.selection_set(cur)

            cal.bind('<<CalendarSelected>>', lambda e: self._pick(cal))
            self._popup.bind('<Escape>', lambda e: self._close())

            self._popup.update_idletasks()
            self._place_popup()
            cal.focus_set()

        def _place_popup(self):
            if not self._popup or not self._popup.winfo_exists():
                return
            pw = self._popup.winfo_reqwidth()
            ph = self._popup.winfo_reqheight()
            sw = self.winfo_screenwidth()
            sh = self.winfo_screenheight()
            wx = self.winfo_rootx()
            wy = self.winfo_rooty() + self.winfo_height()
            x = max(0, min(wx, sw - pw - 4))
            y = wy if wy + ph <= sh else self.winfo_rooty() - ph - 2
            self._popup.geometry(f'+{x}+{max(0, y)}')

        def _pick(self, cal):
            try:
                self.set_date(cal.selection_get())
            except Exception:
                pass
            self._close()
            self.event_generate('<<DateEntrySelected>>')

    CALENDAR_AVAILABLE = True
except ImportError:
    CALENDAR_AVAILABLE = False

import m2r_crm_database as db


class CustomerListPanel(ttk.Frame):
    """Panel containing the customer list with search and filter controls."""

    def __init__(self, parent, on_select: Callable[[int], None]):
        super().__init__(parent)
        self.on_select = on_select
        self._create_widgets()
        self.refresh()

    def _create_widgets(self):
        # Search and filter frame
        controls_frame = ttk.Frame(self)
        controls_frame.pack(fill="x", padx=5, pady=5)

        # New Customer button (branded)
        self.new_btn = ttk.Button(controls_frame, text="+ New Customer", command=self._on_new,
                                  style="Accent.TButton")
        self.new_btn.pack(side="left", padx=(0, 10))

        # Search
        ttk.Label(controls_frame, text="Search:").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *args: self._on_search())
        self.search_entry = ttk.Entry(controls_frame, textvariable=self.search_var, width=25)
        self.search_entry.pack(side="left", padx=(5, 3))

        self.clear_search_btn = ttk.Button(controls_frame, text="Clear", width=5,
                                            command=lambda: self.search_var.set(""))
        self.clear_search_btn.pack(side="left", padx=(0, 10))

        # Filter
        ttk.Label(controls_frame, text="Filter:").pack(side="left")
        self.filter_var = tk.StringVar(value="Active")
        self.filter_combo = ttk.Combobox(
            controls_frame,
            textvariable=self.filter_var,
            values=["Active", "Retired", "All"],
            state="readonly",
            width=10
        )
        self.filter_combo.pack(side="left", padx=5)
        self.filter_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh())

        # Customer list
        list_frame = ttk.Frame(self)
        list_frame.pack(fill="both", expand=True, padx=5, pady=(0, 5))

        # Treeview with scrollbar
        columns = ("account", "company", "contact", "status")
        self.tree = ttk.Treeview(list_frame, columns=columns, show="headings", selectmode="browse")

        self.tree.heading("account", text="Account #", command=lambda: self._sort_by("account_number"))
        self.tree.heading("company", text="Company", command=lambda: self._sort_by("company_name"))
        self.tree.heading("contact", text="Contact", command=lambda: self._sort_by("contact_name"))
        self.tree.heading("status", text="Status", command=lambda: self._sort_by("subscription_status"))

        self.tree.column("account", width=85, minwidth=50, stretch=False)
        self.tree.column("company", width=190, minwidth=80)
        self.tree.column("contact", width=140, minwidth=60)
        self.tree.column("status", width=55, minwidth=30, stretch=False)

        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        self.tree.bind("<Double-1>", self._on_tree_double_click)

        # Count label
        self.count_label = ttk.Label(self, text="0 customers")
        self.count_label.pack(anchor="w", padx=5, pady=(0, 5))

        # Sort state
        self._sort_column = "company_name"
        self._sort_dir = "ASC"

        # For new customer callback
        self._on_new_callback = None

    def set_new_callback(self, callback: Callable[[], None]):
        """Set the callback for New Customer button."""
        self._on_new_callback = callback

    def _on_new(self):
        if self._on_new_callback:
            self._on_new_callback()

    def _on_search(self):
        """Handle search input with debounce."""
        if hasattr(self, '_search_after_id'):
            self.after_cancel(self._search_after_id)
        self._search_after_id = self.after(300, self.refresh)

    def _sort_by(self, column: str):
        """Sort by column, toggling direction."""
        if self._sort_column == column:
            self._sort_dir = "DESC" if self._sort_dir == "ASC" else "ASC"
        else:
            self._sort_column = column
            self._sort_dir = "ASC"
        self.refresh()

    def refresh(self):
        """Refresh the customer list."""
        # Clear existing items
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Get customers
        customers = db.search_customers(
            search_term=self.search_var.get(),
            filter_status=self.filter_var.get(),
            order_by=self._sort_column,
            order_dir=self._sort_dir
        )

        # Populate tree
        for customer in customers:
            status = "Yes" if customer.get('subscription_status') == 'YES' else "No"
            if customer.get('is_retired'):
                status = "Retired"

            self.tree.insert("", "end",
                iid=str(customer['id']),
                values=(
                    customer.get('account_number', ''),
                    customer.get('company_name', ''),
                    customer.get('contact_name', ''),
                    status
                )
            )

        # Update count
        count = len(customers)
        self.count_label.config(text=f"{count} customer{'s' if count != 1 else ''}")

    def _on_tree_select(self, event):
        """Handle tree selection."""
        selection = self.tree.selection()
        if selection:
            customer_id = int(selection[0])
            self.on_select(customer_id)

    def _on_tree_double_click(self, event):
        """Handle double-click on tree item."""
        self._on_tree_select(event)

    def select_customer(self, customer_id: int):
        """Select a customer in the tree by ID."""
        item_id = str(customer_id)
        if self.tree.exists(item_id):
            self.tree.selection_set(item_id)
            self.tree.see(item_id)
            self.tree.focus(item_id)

    def get_selected_id(self) -> Optional[int]:
        """Get the currently selected customer ID."""
        selection = self.tree.selection()
        if selection:
            return int(selection[0])
        return None


class CustomerFormPanel(ttk.Frame):
    """Panel containing the customer detail form."""

    def __init__(self, parent, on_save: Callable[[Dict], None], on_delete: Callable[[int], None]):
        super().__init__(parent)
        self.on_save = on_save
        self.on_delete = on_delete
        self.current_customer_id: Optional[int] = None
        self._saved_form_data: Optional[Dict[str, Any]] = None
        self._original_account_number: Optional[str] = None
        self._create_widgets()

    def _create_widgets(self):
        # Main container with scrollbar
        _bg = ttk.Style().lookup('TFrame', 'background')
        canvas = tk.Canvas(self, highlightthickness=0, bg=_bg)
        scrollbar = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Enable mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        # Configure column weights so license/bureau frames get enough room
        form = self.scrollable_frame
        form.columnconfigure(5, minsize=230)
        row = 0

        # Title and Buttons row
        btn_frame = ttk.Frame(form)
        btn_frame.grid(row=row, column=0, columnspan=7, sticky="w", padx=10, pady=(10, 15))

        ttk.Label(btn_frame, text="Customer Details", font=("Segoe UI", 12, "bold")).pack(side="left")

        self.save_btn = ttk.Button(btn_frame, text="Save", command=self._on_save, width=8,
                                    style="Accent.TButton")
        self.save_btn.pack(side="left", padx=(15, 3))

        self.delete_btn = ttk.Button(btn_frame, text="Delete", command=self._on_delete, width=8)
        self.delete_btn.pack(side="left", padx=3)

        self.cancel_btn = ttk.Button(btn_frame, text="Cancel", command=self._on_cancel, width=8)
        self.cancel_btn.pack(side="left", padx=3)

        self.followup_btn = ttk.Button(btn_frame, text="Follow Up", command=self._on_follow_up, width=10)
        self.followup_btn.pack(side="left", padx=(15, 3))
        row += 1

        # === ROW 1: Account # (left) | AP Email (middle) ===
        ttk.Label(form, text="Account #:").grid(row=row, column=0, sticky="e", padx=(10, 5), pady=3)
        self.account_var = tk.StringVar()
        self.account_entry = ttk.Entry(form, textvariable=self.account_var, width=15)
        self.account_entry.grid(row=row, column=1, sticky="w", pady=3)
        self.account_entry.bind("<FocusOut>", self._on_account_number_change)
        self.auto_btn = ttk.Button(form, text="Auto", width=6, command=self._generate_account)
        self.auto_btn.grid(row=row, column=2, sticky="w", padx=5, pady=3)

        ttk.Label(form, text="AP Email:").grid(row=row, column=3, sticky="e", padx=(15, 5), pady=3)
        self.email_ap_var = tk.StringVar()
        self.email_ap_entry = ttk.Entry(form, textvariable=self.email_ap_var, width=30)
        self.email_ap_entry.grid(row=row, column=4, sticky="w", pady=3)

        # License Info section - far right, spans rows 1-5 (green background)
        _LIC_BG = "#B3CC48"
        _LIC_FG = "#144478"
        license_frame = tk.LabelFrame(form, text="License Info", padx=8, pady=4,
                                      bg=_LIC_BG, fg=_LIC_FG,
                                      font=("Segoe UI", 9, "bold"))
        license_frame.grid(row=row, column=5, rowspan=5, sticky="nsew", padx=(15, 10), pady=3)

        tk.Label(license_frame, text="License:", bg=_LIC_BG, fg=_LIC_FG).grid(row=0, column=0, sticky="e", padx=(0, 5), pady=3)
        self.license_var = tk.StringVar()
        self.license_combo = ttk.Combobox(
            license_frame,
            textvariable=self.license_var,
            values=["Basic Standard", "Basic Network", "Professional", "Enterprise", "Trial", "Demo"],
            width=18
        )
        self.license_combo.grid(row=0, column=1, sticky="w", pady=3)

        tk.Label(license_frame, text="Version:", bg=_LIC_BG, fg=_LIC_FG).grid(row=1, column=0, sticky="e", padx=(0, 5), pady=3)
        self.version_var = tk.StringVar()
        self.version_entry = ttk.Entry(license_frame, textvariable=self.version_var, width=20)
        self.version_entry.grid(row=1, column=1, sticky="w", pady=3)

        tk.Label(license_frame, text="Auth Date:", bg=_LIC_BG, fg=_LIC_FG).grid(row=2, column=0, sticky="e", padx=(0, 5), pady=3)
        if CALENDAR_AVAILABLE:
            self.auth_date_entry = DateEntry(license_frame, width=14, date_pattern='mm/dd/yyyy')
            self.auth_date_entry.delete(0, "end")
        else:
            self.auth_date_var = tk.StringVar()
            self.auth_date_entry = ttk.Entry(license_frame, textvariable=self.auth_date_var, width=15)
        self.auth_date_entry.grid(row=2, column=1, sticky="w", pady=3)

        tk.Label(license_frame, text="Paid Through:", bg=_LIC_BG, fg=_LIC_FG).grid(row=3, column=0, sticky="e", padx=(0, 5), pady=3)
        if CALENDAR_AVAILABLE:
            self.paid_date_entry = DateEntry(license_frame, width=14, date_pattern='mm/dd/yyyy')
            self.paid_date_entry.delete(0, "end")
        else:
            self.paid_date_var = tk.StringVar()
            self.paid_date_entry = ttk.Entry(license_frame, textvariable=self.paid_date_var, width=15)
        self.paid_date_entry.grid(row=3, column=1, sticky="w", pady=3)

        tk.Label(license_frame, text="Support Amt:", bg=_LIC_BG, fg=_LIC_FG).grid(row=4, column=0, sticky="e", padx=(0, 5), pady=3)
        self.support_amt_var = tk.StringVar()
        self.support_amt_entry = ttk.Entry(license_frame, textvariable=self.support_amt_var, width=15)
        self.support_amt_entry.grid(row=4, column=1, sticky="w", pady=3)

        row += 1

        # === ROW 2: Company (left) | User Email (middle) ===
        ttk.Label(form, text="Company:").grid(row=row, column=0, sticky="e", padx=(10, 5), pady=3)
        self.company_var = tk.StringVar()
        self.company_entry = ttk.Entry(form, textvariable=self.company_var, width=30)
        self.company_entry.grid(row=row, column=1, columnspan=2, sticky="w", pady=3)

        ttk.Label(form, text="User Email:").grid(row=row, column=3, sticky="e", padx=(15, 5), pady=3)
        self.email_user_var = tk.StringVar()
        self.email_user_entry = ttk.Entry(form, textvariable=self.email_user_var, width=30)
        self.email_user_entry.grid(row=row, column=4, sticky="w", pady=3)
        row += 1

        # === ROW 3: Contact (left) | IT Email (middle) ===
        ttk.Label(form, text="Contact:").grid(row=row, column=0, sticky="e", padx=(10, 5), pady=3)
        self.contact_var = tk.StringVar()
        self.contact_entry = ttk.Entry(form, textvariable=self.contact_var, width=30)
        self.contact_entry.grid(row=row, column=1, columnspan=2, sticky="w", pady=3)

        ttk.Label(form, text="IT Email:").grid(row=row, column=3, sticky="e", padx=(15, 5), pady=3)
        self.email_it_var = tk.StringVar()
        self.email_it_entry = ttk.Entry(form, textvariable=self.email_it_var, width=30)
        self.email_it_entry.grid(row=row, column=4, sticky="w", pady=3)
        row += 1

        # === ROW 4: Phone (left) ===
        ttk.Label(form, text="Phone:").grid(row=row, column=0, sticky="e", padx=(10, 5), pady=3)
        phone_frame = ttk.Frame(form)
        phone_frame.grid(row=row, column=1, columnspan=2, sticky="w", pady=3)
        self.phone_var = tk.StringVar()
        self._phone_formatting = False  # Flag to prevent recursive formatting
        self.phone_var.trace_add("write", self._format_phone_number)
        self.phone_entry = ttk.Entry(phone_frame, textvariable=self.phone_var, width=15)
        self.phone_entry.pack(side="left")
        ttk.Label(phone_frame, text="Ext:").pack(side="left", padx=(10, 5))
        self.ext_var = tk.StringVar()
        self.ext_entry = ttk.Entry(phone_frame, textvariable=self.ext_var, width=6)
        self.ext_entry.pack(side="left")
        row += 1

        # === ROW 5: Address (left) ===
        ttk.Label(form, text="Address:").grid(row=row, column=0, sticky="e", padx=(10, 5), pady=3)
        self.street_var = tk.StringVar()
        self.street_entry = ttk.Entry(form, textvariable=self.street_var, width=30)
        self.street_entry.grid(row=row, column=1, columnspan=2, sticky="w", pady=3)
        row += 1

        # === ROW 6: City/St/Zip (left) ===
        ttk.Label(form, text="City/St/Zip:").grid(row=row, column=0, sticky="e", padx=(10, 5), pady=3)
        addr_frame = ttk.Frame(form)
        addr_frame.grid(row=row, column=1, columnspan=2, sticky="w", pady=3)
        _US_CA_STATES = [
            "AL", "AK", "AZ", "AR", "CA", "CO", "CT", "DE", "FL", "GA",
            "HI", "ID", "IL", "IN", "IA", "KS", "KY", "LA", "ME", "MD",
            "MA", "MI", "MN", "MS", "MO", "MT", "NE", "NV", "NH", "NJ",
            "NM", "NY", "NC", "ND", "OH", "OK", "OR", "PA", "RI", "SC",
            "SD", "TN", "TX", "UT", "VT", "VA", "WA", "WV", "WI", "WY",
            "DC",
            "AB", "BC", "MB", "NB", "NL", "NS", "NT", "NU", "ON", "PE",
            "QC", "SK", "YT",
        ]
        self.city_var = tk.StringVar()
        self.city_entry = ttk.Entry(addr_frame, textvariable=self.city_var, width=15)
        self.city_entry.pack(side="left")
        self.state_var = tk.StringVar()
        self.state_combo = ttk.Combobox(addr_frame, textvariable=self.state_var,
                                        values=_US_CA_STATES, width=4, state="normal")
        self.state_combo.pack(side="left", padx=(5, 0))
        self.zip_var = tk.StringVar()
        self.zip_entry = ttk.Entry(addr_frame, textvariable=self.zip_var, width=10)
        self.zip_entry.pack(side="left", padx=(5, 0))

        # Bureau IDs section - far right, spans rows 6-10 (same height as License Info)
        bureau_frame = ttk.LabelFrame(form, text="Bureau IDs", padding=8)
        bureau_frame.grid(row=row, column=5, rowspan=5, sticky="nsew", padx=(15, 10), pady=3)

        ttk.Label(bureau_frame, text="Company ID:").grid(row=0, column=0, sticky="e", padx=(0, 5), pady=3)
        self.company_id_var = tk.StringVar()
        self.company_id_entry = ttk.Entry(bureau_frame, textvariable=self.company_id_var, width=18)
        self.company_id_entry.grid(row=0, column=1, sticky="w", pady=3)

        ttk.Label(bureau_frame, text="Equifax:").grid(row=1, column=0, sticky="e", padx=(0, 5), pady=3)
        self.bureau_equifax_var = tk.StringVar()
        self.bureau_equifax_entry = ttk.Entry(bureau_frame, textvariable=self.bureau_equifax_var, width=18)
        self.bureau_equifax_entry.grid(row=1, column=1, sticky="w", pady=3)

        ttk.Label(bureau_frame, text="Experian:").grid(row=2, column=0, sticky="e", padx=(0, 5), pady=3)
        self.bureau_experian_var = tk.StringVar()
        self.bureau_experian_entry = ttk.Entry(bureau_frame, textvariable=self.bureau_experian_var, width=18)
        self.bureau_experian_entry.grid(row=2, column=1, sticky="w", pady=3)

        ttk.Label(bureau_frame, text="TransUnion:").grid(row=3, column=0, sticky="e", padx=(0, 5), pady=3)
        self.bureau_transunion_var = tk.StringVar()
        self.bureau_transunion_entry = ttk.Entry(bureau_frame, textvariable=self.bureau_transunion_var, width=18)
        self.bureau_transunion_entry.grid(row=3, column=1, sticky="w", pady=3)

        row += 1

        # === ROW 7-9: empty left rows for bureau section to span ===
        row += 1
        row += 1
        row += 1

        # === ROW 10: Status + Retired + Checkboxes ===
        status_frame = ttk.Frame(form)
        status_frame.grid(row=row, column=0, columnspan=6, sticky="w", padx=(10, 0), pady=3)
        ttk.Label(status_frame, text="Status:").pack(side="left", padx=(0, 5))
        self.status_var = tk.StringVar(value="NO")
        self.status_combo = ttk.Combobox(
            status_frame,
            textvariable=self.status_var,
            values=["YES", "NO"],
            state="readonly",
            width=6
        )
        self.status_combo.pack(side="left")
        self.retired_var = tk.BooleanVar()
        self.retired_check = ttk.Checkbutton(status_frame, text="Retired", variable=self.retired_var)
        self.retired_check.pack(side="left", padx=(15, 0))
        self.no_support_var = tk.BooleanVar()
        ttk.Checkbutton(status_frame, text="No Support", variable=self.no_support_var).pack(side="left", padx=(15, 0))
        self.support_inv_var = tk.BooleanVar()
        ttk.Checkbutton(status_frame, text="Support Invoice", variable=self.support_inv_var).pack(side="left", padx=(10, 0))
        self.past_due_inv_var = tk.BooleanVar()
        ttk.Checkbutton(status_frame, text="Past Due Invoice", variable=self.past_due_inv_var).pack(side="left", padx=(10, 0))
        self.no_inv_var = tk.BooleanVar()
        ttk.Checkbutton(status_frame, text="No Invoice", variable=self.no_inv_var).pack(side="left", padx=(10, 0))
        row += 1

        # Invoice dates (read-only display)
        ttk.Label(form, text="Last Invoice:").grid(row=row, column=0, sticky="e", padx=(10, 5), pady=3)
        self.last_inv_label = ttk.Label(form, text="-")
        self.last_inv_label.grid(row=row, column=1, sticky="w", pady=3)

        ttk.Label(form, text="Last Past Due:").grid(row=row, column=2, sticky="e", padx=(10, 5), pady=3)
        self.last_past_due_label = ttk.Label(form, text="-")
        self.last_past_due_label.grid(row=row, column=3, sticky="w", pady=3)
        row += 1

        # Separator
        ttk.Separator(form, orient="horizontal").grid(row=row, column=0, columnspan=6, sticky="ew", pady=10, padx=10)
        row += 1

        # Note to Customer
        ttk.Label(form, text="Note to Customer:").grid(row=row, column=0, sticky="ne", padx=(10, 5), pady=3)
        self.note_customer_text = tk.Text(form, width=60, height=3)
        self.note_customer_text.grid(row=row, column=1, columnspan=4, sticky="w", pady=3)
        row += 1

        # Memo to Self
        ttk.Label(form, text="Memo to Self:").grid(row=row, column=0, sticky="ne", padx=(10, 5), pady=3)
        self.memo_text = tk.Text(form, width=60, height=3)
        self.memo_text.grid(row=row, column=1, columnspan=4, sticky="w", pady=3)

    def _generate_account(self):
        """Generate next account number."""
        # Check if this is an existing customer
        if self.current_customer_id and self._original_account_number:
            next_num = db.get_next_account_number()
            result = messagebox.askyesno(
                "Change Account Number",
                f"You are changing the account number from:\n"
                f"  {self._original_account_number}\n"
                f"to:\n"
                f"  {next_num}\n\n"
                f"Do you want to keep this change?",
                icon="warning"
            )
            if not result:
                return  # User cancelled, don't change

        next_num = db.get_next_account_number()
        self.account_var.set(next_num)

    def _on_account_number_change(self, event=None):
        """Check if account number was changed on an existing customer."""
        # Only check for existing customers
        if not self.current_customer_id or not self._original_account_number:
            return

        current_value = self.account_var.get().strip()

        # If unchanged, nothing to do
        if current_value == self._original_account_number:
            return

        # Warn user about changing account number
        result = messagebox.askyesno(
            "Change Account Number",
            f"You are changing the account number from:\n"
            f"  {self._original_account_number}\n"
            f"to:\n"
            f"  {current_value}\n\n"
            f"Do you want to keep this change?",
            icon="warning"
        )

        if not result:
            # Revert to original account number
            self.account_var.set(self._original_account_number)

    def _format_phone_number(self, *args):
        """Auto-format phone number as (XXX)XXX-XXXX."""
        if self._phone_formatting:
            return

        self._phone_formatting = True
        try:
            # Get current value and cursor position
            current = self.phone_var.get()

            # Extract only digits
            digits = ''.join(c for c in current if c.isdigit())

            # Limit to 10 digits
            digits = digits[:10]

            # Format based on number of digits
            if len(digits) == 0:
                formatted = ""
            elif len(digits) <= 3:
                formatted = f"({digits}"
            elif len(digits) <= 6:
                formatted = f"({digits[:3]}){digits[3:]}"
            else:
                formatted = f"({digits[:3]}){digits[3:6]}-{digits[6:]}"

            # Only update if different to avoid cursor issues
            if current != formatted:
                self.phone_var.set(formatted)
                # Move cursor to end
                self.phone_entry.icursor(len(formatted))
        finally:
            self._phone_formatting = False

    def load_customer(self, customer_id: int):
        """Load a customer into the form."""
        customer = db.get_customer(customer_id)
        if not customer:
            return

        self.clear_form()
        self.current_customer_id = customer_id

        # Basic info
        self.account_var.set(customer.get('account_number', ''))
        self._original_account_number = customer.get('account_number', '')
        self.company_var.set(customer.get('company_name', ''))
        self.contact_var.set(customer.get('contact_name', ''))
        self.email_ap_var.set(customer.get('email_ap', '') or customer.get('email', ''))
        self.email_user_var.set(customer.get('email_user', '') or '')
        self.email_it_var.set(customer.get('email_it', '') or '')
        self.phone_var.set(customer.get('phone', ''))
        self.ext_var.set(customer.get('phone_extension', ''))

        # Address
        self.street_var.set(customer.get('address_street', ''))
        self.city_var.set(customer.get('address_city', ''))
        self.state_var.set(customer.get('address_state', ''))
        self.zip_var.set(customer.get('address_zip', ''))

        # License info
        self.license_var.set(customer.get('license_type', ''))
        self.version_var.set(customer.get('software_version', ''))
        self.company_id_var.set(customer.get('company_id', '') or '')
        self.bureau_equifax_var.set(customer.get('bureau_equifax', '') or customer.get('bureau_id', '') or '')
        self.bureau_experian_var.set(customer.get('bureau_experian', '') or '')
        self.bureau_transunion_var.set(customer.get('bureau_transunion', '') or '')

        # Dates
        self._set_date_entry(self.auth_date_entry, customer.get('auth_date'))
        self._set_date_entry(self.paid_date_entry, customer.get('paid_through_date'))

        # Financial
        amt = customer.get('support_amount')
        self.support_amt_var.set(f"{amt:.2f}" if amt else "")

        # Status
        self.status_var.set(customer.get('subscription_status', 'NO'))
        self.retired_var.set(bool(customer.get('is_retired', 0)))
        self.no_support_var.set(bool(customer.get('no_support', 0)))
        self.support_inv_var.set(bool(customer.get('support_invoice', 0)))
        self.past_due_inv_var.set(bool(customer.get('past_due_invoice', 0)))
        self.no_inv_var.set(bool(customer.get('no_invoice', 0)))

        # Notes
        self.note_customer_text.delete("1.0", "end")
        self.note_customer_text.insert("1.0", customer.get('note_to_customer', '') or '')

        self.memo_text.delete("1.0", "end")
        self.memo_text.insert("1.0", customer.get('memo_to_self', '') or '')

        # Invoice dates (display only)
        self._format_date_label(self.last_inv_label, customer.get('last_invoice_date'))
        self._format_date_label(self.last_past_due_label, customer.get('last_invoice_past_due'))

        # Enable delete button
        self.delete_btn.config(state="normal")

        # Store form state for change detection
        self._store_form_state()

    def _set_date_entry(self, entry, date_value):
        """Set a date entry value."""
        if CALENDAR_AVAILABLE:
            entry.delete(0, "end")
            if date_value:
                try:
                    if isinstance(date_value, str):
                        dt = datetime.fromisoformat(date_value)
                        entry.set_date(dt.date())
                    elif isinstance(date_value, date):
                        entry.set_date(date_value)
                except:
                    pass
        else:
            var = getattr(entry, 'textvariable', None)
            if var:
                if date_value:
                    try:
                        if isinstance(date_value, str):
                            dt = datetime.fromisoformat(date_value)
                            entry.delete(0, "end")
                            entry.insert(0, dt.strftime("%m/%d/%Y"))
                        elif isinstance(date_value, date):
                            entry.delete(0, "end")
                            entry.insert(0, date_value.strftime("%m/%d/%Y"))
                    except:
                        pass

    def _format_date_label(self, label, date_value):
        """Format a date for display in a label."""
        if date_value:
            try:
                if isinstance(date_value, str):
                    dt = datetime.fromisoformat(date_value)
                    label.config(text=dt.strftime("%m/%d/%Y"))
                elif isinstance(date_value, date):
                    label.config(text=date_value.strftime("%m/%d/%Y"))
                else:
                    label.config(text="-")
            except:
                label.config(text="-")
        else:
            label.config(text="-")

    def _get_date_value(self, entry) -> Optional[str]:
        """Get date value from entry as ISO format string."""
        if CALENDAR_AVAILABLE:
            try:
                date_str = entry.get()
                if not date_str:
                    return None
                dt = entry.get_date()
                return dt.isoformat()
            except:
                return None
        else:
            date_str = entry.get().strip()
            if not date_str:
                return None
            for fmt in ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d']:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.date().isoformat()
                except ValueError:
                    continue
            return None

    def clear_form(self):
        """Clear all form fields."""
        self.current_customer_id = None
        self._saved_form_data = None
        self._original_account_number = None

        self.account_var.set("")
        self.company_var.set("")
        self.contact_var.set("")
        self.email_ap_var.set("")
        self.email_user_var.set("")
        self.email_it_var.set("")
        self.phone_var.set("")
        self.ext_var.set("")
        self.street_var.set("")
        self.city_var.set("")
        self.state_var.set("")
        self.zip_var.set("")
        self.license_var.set("")
        self.version_var.set("")
        self.company_id_var.set("")
        self.bureau_equifax_var.set("")
        self.bureau_experian_var.set("")
        self.bureau_transunion_var.set("")
        self.support_amt_var.set("")
        self.status_var.set("NO")
        self.retired_var.set(False)
        self.no_support_var.set(False)
        self.support_inv_var.set(False)
        self.past_due_inv_var.set(False)
        self.no_inv_var.set(False)

        if CALENDAR_AVAILABLE:
            self.auth_date_entry.delete(0, "end")
            self.paid_date_entry.delete(0, "end")
        else:
            self.auth_date_entry.delete(0, "end")
            self.paid_date_entry.delete(0, "end")

        self.note_customer_text.delete("1.0", "end")
        self.memo_text.delete("1.0", "end")

        self.last_inv_label.config(text="-")
        self.last_past_due_label.config(text="-")

        # Disable delete for new customers
        self.delete_btn.config(state="disabled")

    def new_customer(self):
        """Prepare form for a new customer."""
        self.clear_form()
        self._generate_account()
        self.company_entry.focus()
        # Store form state for change detection
        self._store_form_state()

    def get_form_data(self) -> Dict[str, Any]:
        """Get all form data as a dictionary."""
        data = {
            'account_number': self.account_var.get().strip(),
            'company_name': self.company_var.get().strip(),
            'contact_name': self.contact_var.get().strip(),
            'email_ap': self.email_ap_var.get().strip(),
            'email_user': self.email_user_var.get().strip(),
            'email_it': self.email_it_var.get().strip(),
            'phone': self.phone_var.get().strip(),
            'phone_extension': self.ext_var.get().strip(),
            'address_street': self.street_var.get().strip(),
            'address_city': self.city_var.get().strip(),
            'address_state': self.state_var.get().strip(),
            'address_zip': self.zip_var.get().strip(),
            'license_type': self.license_var.get().strip(),
            'software_version': self.version_var.get().strip(),
            'company_id': self.company_id_var.get().strip(),
            'bureau_equifax': self.bureau_equifax_var.get().strip(),
            'bureau_experian': self.bureau_experian_var.get().strip(),
            'bureau_transunion': self.bureau_transunion_var.get().strip(),
            'auth_date': self._get_date_value(self.auth_date_entry),
            'paid_through_date': self._get_date_value(self.paid_date_entry),
            'subscription_status': self.status_var.get(),
            'is_retired': 1 if self.retired_var.get() else 0,
            'no_support': 1 if self.no_support_var.get() else 0,
            'support_invoice': 1 if self.support_inv_var.get() else 0,
            'past_due_invoice': 1 if self.past_due_inv_var.get() else 0,
            'no_invoice': 1 if self.no_inv_var.get() else 0,
            'note_to_customer': self.note_customer_text.get("1.0", "end-1c").strip(),
            'memo_to_self': self.memo_text.get("1.0", "end-1c").strip(),
        }

        # Parse support amount
        amt_str = self.support_amt_var.get().strip()
        if amt_str:
            try:
                data['support_amount'] = float(amt_str.replace(',', ''))
            except ValueError:
                data['support_amount'] = None
        else:
            data['support_amount'] = None

        return data

    def _store_form_state(self):
        """Store the current form state for change detection."""
        self._saved_form_data = self.get_form_data()

    def has_unsaved_changes(self) -> bool:
        """Check if the form has unsaved changes."""
        if self._saved_form_data is None:
            return False
        current_data = self.get_form_data()
        return current_data != self._saved_form_data

    def check_unsaved_changes(self) -> bool:
        """
        Check for unsaved changes and prompt user if needed.
        Returns True if OK to proceed, False if user wants to stay.
        """
        if not self.has_unsaved_changes():
            return True

        result = messagebox.askyesnocancel(
            "Unsaved Changes",
            "You have unsaved changes. Do you want to save before continuing?",
            icon="warning"
        )

        if result is None:  # Cancel
            return False
        elif result:  # Yes - save first
            self._on_save()
            return True
        else:  # No - discard changes
            return True

    def _on_save(self):
        """Handle save button click."""
        data = self.get_form_data()

        # Validate required fields
        if not data['account_number']:
            messagebox.showerror("Validation Error", "Account number is required.")
            self.account_entry.focus()
            return

        # Check account number uniqueness
        if not db.is_account_number_unique(data['account_number'], self.current_customer_id):
            messagebox.showerror("Validation Error", "Account number already exists.")
            self.account_entry.focus()
            return

        self.on_save(data)

    def _on_delete(self):
        """Handle delete button click."""
        if self.current_customer_id:
            self.on_delete(self.current_customer_id)

    def _on_cancel(self):
        """Handle cancel button click."""
        if self.current_customer_id:
            self.load_customer(self.current_customer_id)
        else:
            self.clear_form()

    def _on_follow_up(self):
        """Open the follow-up dialog for the current customer."""
        if not self.current_customer_id:
            messagebox.showwarning("No Customer", "Please select or save a customer first.")
            return
        company = self.company_var.get().strip() or "Customer"
        dialog = FollowUpDialog(self.winfo_toplevel(), self.current_customer_id, company)
        self.wait_window(dialog)

class NotesPanel(ttk.Frame):
    """Panel for displaying and adding customer notes."""

    def __init__(self, parent):
        super().__init__(parent)
        self.current_customer_id: Optional[int] = None
        self._create_widgets()

    def _create_widgets(self):
        # Title
        title_frame = ttk.Frame(self)
        title_frame.pack(fill="x", padx=5, pady=(10, 5))

        ttk.Label(title_frame, text="Notes", font=("Segoe UI", 11, "bold")).pack(side="left")

        # Notes list
        list_frame = ttk.Frame(self)
        list_frame.pack(fill="both", expand=True, padx=5, pady=5)

        self.notes_text = tk.Text(list_frame, height=8, state="disabled", wrap="word")
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.notes_text.yview)
        self.notes_text.configure(yscrollcommand=scrollbar.set)

        self.notes_text.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Common notes dropdown
        common_frame = ttk.Frame(self)
        common_frame.pack(fill="x", padx=5, pady=(0, 3))

        ttk.Label(common_frame, text="Common:").pack(side="left")
        self.common_notes = [
            "Received order via web. Sent onboarding email to customer.",
            "Received onboarding registration from customer. Sent Activation Key.",
            "Received annual renewal. Updated paid through. Sent download reminder to customer.",
            "Annual renewal invoice mailed to customer.",
            "Recurring annual renewal invoice created.",
            "Customer requested cancellation via email/phone. No longer using M2R.",
            "Annual renewal not paid. Retired account.",
        ]
        self.common_note_var = tk.StringVar()
        self.common_note_combo = ttk.Combobox(
            common_frame,
            textvariable=self.common_note_var,
            values=self.common_notes,
            state="readonly",
            width=70
        )
        self.common_note_combo.pack(side="left", padx=(5, 0), fill="x", expand=True)
        self.common_note_combo.bind("<<ComboboxSelected>>", self._on_common_note_selected)

        # Add note frame
        add_frame = ttk.Frame(self)
        add_frame.pack(fill="x", padx=5, pady=(0, 10))

        self.new_note_var = tk.StringVar()
        self.new_note_entry = ttk.Entry(add_frame, textvariable=self.new_note_var)
        self.new_note_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.new_note_entry.bind("<Return>", lambda e: self._add_note())

        self.add_btn = ttk.Button(add_frame, text="Add Note", command=self._add_note)
        self.add_btn.pack(side="left")

    def load_notes(self, customer_id: int):
        """Load notes for a customer."""
        self.current_customer_id = customer_id
        notes = db.get_notes(customer_id)

        self.notes_text.config(state="normal")
        self.notes_text.delete("1.0", "end")

        for note in notes:
            date_str = note.get('note_date', '')
            if date_str:
                try:
                    dt = datetime.fromisoformat(date_str)
                    date_str = dt.strftime("%m/%d/%Y %I:%M %p")
                except:
                    pass

            text = note.get('note_text', '')
            self.notes_text.insert("end", f"{date_str}\n{text}\n\n")

        self.notes_text.config(state="disabled")

        # Enable/disable add note
        self.add_btn.config(state="normal")
        self.new_note_entry.config(state="normal")

    def clear_notes(self):
        """Clear the notes display."""
        self.current_customer_id = None
        self.notes_text.config(state="normal")
        self.notes_text.delete("1.0", "end")
        self.notes_text.config(state="disabled")
        self.new_note_var.set("")
        self.add_btn.config(state="disabled")
        self.new_note_entry.config(state="disabled")

    def _add_note(self):
        """Add a new note."""
        if not self.current_customer_id:
            return

        note_text = self.new_note_var.get().strip()
        if not note_text:
            return

        db.create_note(self.current_customer_id, note_text)
        self.new_note_var.set("")
        self.common_note_combo.set("")
        self.load_notes(self.current_customer_id)

    def _on_common_note_selected(self, event):
        """Populate the note entry with the selected common note."""
        selected = self.common_note_var.get()
        if selected:
            self.new_note_var.set(selected)


class FollowUpDialog(tk.Toplevel):
    """Dialog for viewing and adding follow-ups for a specific customer."""

    def __init__(self, parent, customer_id: int, customer_name: str):
        super().__init__(parent)
        self.customer_id = customer_id
        self.title(f"Follow-ups — {customer_name}")
        self.geometry("680x500")
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()

        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - 680) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 500) // 2
        self.geometry(f"+{x}+{y}")

        self._create_widgets()
        self._load()

    def _create_widgets(self):
        # Existing follow-ups list
        list_frame = ttk.LabelFrame(self, text="Existing Follow-ups", padding=8)
        list_frame.pack(fill="both", expand=True, padx=10, pady=(10, 5))

        cols = ('created', 'due', 'note', 'status')
        self.tree = ttk.Treeview(list_frame, columns=cols, show='headings', height=8)
        self.tree.heading('created', text='Created')
        self.tree.heading('due', text='Due Date')
        self.tree.heading('note', text='Follow-up Note')
        self.tree.heading('status', text='Status')
        self.tree.column('created', width=105, stretch=False)
        self.tree.column('due', width=85, stretch=False)
        self.tree.column('note', width=360)
        self.tree.column('status', width=75, stretch=False)

        sb = ttk.Scrollbar(list_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        self.tree.pack(side='left', fill='both', expand=True)
        sb.pack(side='right', fill='y')

        self.tree.tag_configure('complete', foreground='gray')

        # Action buttons
        action_frame = ttk.Frame(self)
        action_frame.pack(fill='x', padx=10, pady=(0, 5))
        ttk.Button(action_frame, text="Mark Complete", command=self._mark_complete).pack(side='left', padx=(0, 5))
        ttk.Button(action_frame, text="Reopen", command=self._reopen).pack(side='left', padx=(0, 5))
        ttk.Button(action_frame, text="Delete", command=self._delete_selected).pack(side='left')

        # Add new follow-up frame
        add_frame = ttk.LabelFrame(self, text="Add New Follow-up", padding=8)
        add_frame.pack(fill='x', padx=10, pady=(0, 5))

        ttk.Label(add_frame, text="Note:").grid(row=0, column=0, sticky='e', padx=(0, 5), pady=3)
        self.note_var = tk.StringVar()
        self.note_entry = ttk.Entry(add_frame, textvariable=self.note_var, width=55)
        self.note_entry.grid(row=0, column=1, sticky='ew', pady=3)
        self.note_entry.bind('<Return>', lambda e: self._add())

        ttk.Label(add_frame, text="Due Date:").grid(row=1, column=0, sticky='e', padx=(0, 5), pady=3)
        due_frame = ttk.Frame(add_frame)
        due_frame.grid(row=1, column=1, sticky='w', pady=3)

        if CALENDAR_AVAILABLE:
            self.due_entry = DateEntry(due_frame, width=12)
            self.due_entry.delete(0, 'end')
            self.due_entry.pack(side='left')
        else:
            self.due_var = tk.StringVar()
            self.due_entry = ttk.Entry(due_frame, textvariable=self.due_var, width=12)
            self.due_entry.pack(side='left')
        ttk.Button(due_frame, text="Clear", command=self._clear_due).pack(side='left', padx=(5, 0))

        ttk.Button(add_frame, text="Add Follow-up", command=self._add,
                   style="Accent.TButton").grid(row=2, column=1, sticky='w', pady=(5, 0))
        add_frame.columnconfigure(1, weight=1)

        # Close
        ttk.Button(self, text="Close", command=self.destroy).pack(pady=(2, 10))

    def _clear_due(self):
        if CALENDAR_AVAILABLE:
            self.due_entry.delete(0, 'end')
        else:
            self.due_var.set('')

    def _get_due(self) -> Optional[str]:
        if CALENDAR_AVAILABLE:
            try:
                d = self.due_entry.get_date()
                return d.isoformat() if d else None
            except Exception:
                return None
        else:
            val = self.due_var.get().strip()
            if not val:
                return None
            for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(val, fmt).date().isoformat()
                except ValueError:
                    pass
            return None

    def _load(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for fup in db.get_follow_ups(self.customer_id):
            created = _fmt_dt(fup.get('created_date'))
            due = _fmt_dt(fup.get('due_date')) if fup.get('due_date') else ''
            status = fup.get('status', 'Open')
            tag = 'complete' if status == 'Complete' else ''
            self.tree.insert('', 'end', iid=str(fup['id']),
                             values=(created, due, fup.get('note_text', ''), status),
                             tags=(tag,))

    def _add(self):
        note = self.note_var.get().strip()
        if not note:
            messagebox.showwarning("Required", "Please enter a follow-up note.", parent=self)
            return
        db.create_follow_up(self.customer_id, note, self._get_due())
        self.note_var.set('')
        self._clear_due()
        self._load()

    def _mark_complete(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Item", "Please select a follow-up first.", parent=self)
            return
        db.update_follow_up_status(int(sel[0]), 'Complete')
        self._load()

    def _reopen(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Item", "Please select a follow-up first.", parent=self)
            return
        db.update_follow_up_status(int(sel[0]), 'Open')
        self._load()

    def _delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Item", "Please select a follow-up first.", parent=self)
            return
        if messagebox.askyesno("Confirm Delete", "Delete this follow-up?", parent=self):
            db.delete_follow_up(int(sel[0]))
            self._load()


def _fmt_dt(val) -> str:
    """Format an ISO datetime/date string as mm/dd/yyyy."""
    if not val:
        return ''
    try:
        return datetime.fromisoformat(str(val)).strftime('%m/%d/%Y')
    except Exception:
        return str(val)


class FollowUpReportWindow(tk.Toplevel):
    """Report window showing all follow-ups with status editing and export."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Follow-up Report")
        self.geometry("950x600")
        self.resizable(True, True)
        self.transient(parent)

        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - 950) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 600) // 2
        self.geometry(f"+{x}+{y}")

        self._status_combo = None  # inline editor widget
        self._editing_iid = None
        self._create_widgets()
        self._load()

    def _create_widgets(self):
        # Toolbar
        toolbar = ttk.Frame(self)
        toolbar.pack(fill='x', padx=10, pady=(10, 5))

        ttk.Label(toolbar, text="Show:").pack(side='left')
        self.filter_var = tk.StringVar(value='Open')
        filter_combo = ttk.Combobox(toolbar, textvariable=self.filter_var,
                                    values=['Open', 'Complete', 'All'],
                                    state='readonly', width=10)
        filter_combo.pack(side='left', padx=(5, 15))
        filter_combo.bind('<<ComboboxSelected>>', lambda e: self._load())

        ttk.Button(toolbar, text="Refresh", command=self._load).pack(side='left', padx=(0, 5))
        ttk.Button(toolbar, text="Mark Complete", command=self._mark_complete,
                   style="Accent.TButton").pack(side='left', padx=(0, 5))
        ttk.Button(toolbar, text="Reopen", command=self._reopen).pack(side='left', padx=(0, 5))
        ttk.Button(toolbar, text="Export...", command=self._export).pack(side='left', padx=(15, 0))

        self.count_label = ttk.Label(toolbar, text="")
        self.count_label.pack(side='right')

        # Treeview
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill='both', expand=True, padx=10, pady=(0, 5))

        cols = ('account', 'company', 'contact', 'created', 'due', 'note', 'status')
        self.tree = ttk.Treeview(tree_frame, columns=cols, show='headings')
        self.tree.heading('account', text='Account #')
        self.tree.heading('company', text='Company')
        self.tree.heading('contact', text='Contact')
        self.tree.heading('created', text='Created')
        self.tree.heading('due', text='Due Date')
        self.tree.heading('note', text='Follow-up Note')
        self.tree.heading('status', text='Status')

        self.tree.column('account', width=90, stretch=False)
        self.tree.column('company', width=160)
        self.tree.column('contact', width=120)
        self.tree.column('created', width=90, stretch=False)
        self.tree.column('due', width=85, stretch=False)
        self.tree.column('note', width=280)
        self.tree.column('status', width=80, stretch=False)

        vsb = ttk.Scrollbar(tree_frame, orient='vertical', command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient='horizontal', command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree.tag_configure('complete', foreground='gray')
        self.tree.tag_configure('overdue', foreground='red')

        # Click on Status column to edit inline
        self.tree.bind('<ButtonRelease-1>', self._on_tree_click)

        # Close button
        ttk.Button(self, text="Close", command=self.destroy).pack(pady=(0, 10))

    def _load(self):
        self._close_inline_editor()
        for item in self.tree.get_children():
            self.tree.delete(item)

        rows = db.get_all_follow_ups(self.filter_var.get())
        today = date.today()

        for fup in rows:
            due_str = ''
            tag = ''
            if fup.get('due_date'):
                due_str = _fmt_dt(fup['due_date'])
                try:
                    due_d = date.fromisoformat(str(fup['due_date'])[:10])
                    if fup.get('status') == 'Open' and due_d < today:
                        tag = 'overdue'
                except Exception:
                    pass

            status = fup.get('status', 'Open')
            if tag != 'overdue':
                tag = 'complete' if status == 'Complete' else ''

            self.tree.insert('', 'end', iid=str(fup['id']),
                             values=(
                                 fup.get('account_number', ''),
                                 fup.get('company_name', ''),
                                 fup.get('contact_name', ''),
                                 _fmt_dt(fup.get('created_date')),
                                 due_str,
                                 fup.get('note_text', ''),
                                 status,
                             ),
                             tags=(tag,))

        count = len(rows)
        self.count_label.config(text=f"{count} item{'s' if count != 1 else ''}")

    def _on_tree_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != 'cell':
            self._close_inline_editor()
            return
        col = self.tree.identify_column(event.x)
        iid = self.tree.identify_row(event.y)
        if not iid:
            self._close_inline_editor()
            return
        # col '#7' is the status column (1-indexed)
        if col == '#7':
            self._open_inline_editor(iid)
        else:
            self._close_inline_editor()

    def _open_inline_editor(self, iid):
        self._close_inline_editor()
        bbox = self.tree.bbox(iid, '#7')
        if not bbox:
            return
        x, y, w, h = bbox
        current_status = self.tree.set(iid, 'status')
        var = tk.StringVar(value=current_status)
        combo = ttk.Combobox(self.tree, textvariable=var,
                             values=['Open', 'Complete'],
                             state='readonly', width=9)
        combo.place(x=x, y=y, width=w, height=h)
        combo.focus_set()
        self._status_combo = combo
        self._editing_iid = iid

        def on_select(e=None):
            new_status = var.get()
            db.update_follow_up_status(int(iid), new_status)
            self._close_inline_editor()
            self._load()

        def on_focusout(e=None):
            self._close_inline_editor()

        combo.bind('<<ComboboxSelected>>', on_select)
        combo.bind('<FocusOut>', on_focusout)
        combo.bind('<Escape>', lambda e: self._close_inline_editor())

    def _close_inline_editor(self):
        if self._status_combo:
            try:
                self._status_combo.destroy()
            except Exception:
                pass
            self._status_combo = None
            self._editing_iid = None

    def _mark_complete(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Item", "Please select a follow-up first.", parent=self)
            return
        for iid in sel:
            db.update_follow_up_status(int(iid), 'Complete')
        self._load()

    def _reopen(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showwarning("Select Item", "Please select a follow-up first.", parent=self)
            return
        for iid in sel:
            db.update_follow_up_status(int(iid), 'Open')
        self._load()

    def _export(self):
        rows = db.get_all_follow_ups(self.filter_var.get())
        if not rows:
            messagebox.showwarning("No Data", "No follow-ups to export.", parent=self)
            return

        from tkinter import filedialog
        import os
        default_name = f"FollowUp_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save Report",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("CSV File", "*.csv")],
            initialfile=default_name,
            parent=self
        )
        if not path:
            return

        try:
            if path.lower().endswith('.csv'):
                self._export_csv(rows, path)
            else:
                self._export_excel(rows, path)
            messagebox.showinfo("Export Complete", f"Report exported to:\n{path}", parent=self)
            if messagebox.askyesno("Open File?", "Open the exported file?", parent=self):
                os.startfile(path)
        except Exception as e:
            messagebox.showerror("Export Error", str(e), parent=self)

    def _export_csv(self, rows, path):
        import csv
        headers = ['Account #', 'Company', 'Contact', 'Created', 'Due Date', 'Follow-up Note', 'Status', 'Completed Date']
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for r in rows:
                writer.writerow([
                    r.get('account_number', ''),
                    r.get('company_name', ''),
                    r.get('contact_name', ''),
                    _fmt_dt(r.get('created_date')),
                    _fmt_dt(r.get('due_date')) if r.get('due_date') else '',
                    r.get('note_text', ''),
                    r.get('status', ''),
                    _fmt_dt(r.get('completed_date')) if r.get('completed_date') else '',
                ])

    def _export_excel(self, rows, path):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            # Fall back to CSV
            self._export_csv(rows, path.replace('.xlsx', '.csv'))
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Follow-up Report"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="144478", end_color="144478", fill_type="solid")
        thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['Account #', 'Company', 'Contact', 'Created', 'Due Date',
                   'Follow-up Note', 'Status', 'Completed Date']
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin

        for ri, r in enumerate(rows, 2):
            vals = [
                r.get('account_number', ''),
                r.get('company_name', ''),
                r.get('contact_name', ''),
                _fmt_dt(r.get('created_date')),
                _fmt_dt(r.get('due_date')) if r.get('due_date') else '',
                r.get('note_text', ''),
                r.get('status', ''),
                _fmt_dt(r.get('completed_date')) if r.get('completed_date') else '',
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.border = thin

        # Auto-width
        for ci, header in enumerate(headers, 1):
            max_len = len(header)
            for ri in range(2, len(rows) + 2):
                v = ws.cell(row=ri, column=ci).value
                if v:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)

        ws.freeze_panes = "A2"

        # Meta sheet
        meta = wb.create_sheet("Export Info")
        meta.cell(row=1, column=1, value="Export Date:").font = Font(bold=True)
        meta.cell(row=1, column=2, value=datetime.now().strftime("%m/%d/%Y %I:%M %p"))
        meta.cell(row=2, column=1, value="Filter:").font = Font(bold=True)
        meta.cell(row=2, column=2, value=self.filter_var.get())
        meta.cell(row=3, column=1, value="Total Records:").font = Font(bold=True)
        meta.cell(row=3, column=2, value=len(rows))
        meta.column_dimensions['A'].width = 18
        meta.column_dimensions['B'].width = 30

        wb.save(path)


class EmailCampaignDialog(tk.Toplevel):
    """Dialog for selecting an audience, previewing, and sending an M2R email campaign."""

    _BLUE  = "#144478"
    _WHITE = "#FFFFFF"

    # Maps display label → internal audience key passed to get_send_list()
    _AUDIENCE_OPTIONS = {
        "All (active + recently retired)": "all",
        "Active only":                     "active",
        "Recently retired only":           "retired",
    }

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Send Email Campaign")
        self.geometry("730x680")
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width()  - 730) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 660) // 2
        self.geometry(f"+{x}+{max(0, y)}")

        self._records: list           = []
        self._tagged: set             = set()
        self._preview_done: bool      = False
        self._sending: bool           = False
        self._sent_ok: int            = 0
        self._sent_fail: int          = 0
        self._sent_skipped: int       = 0
        self._log_path: Optional[str] = None
        _cfg                          = _load_config()
        self._last_folder: str        = _cfg.get("newsletter_last_folder", "")

        self.tagged_only_var = tk.BooleanVar(value=False)
        self.tagged_only_var.trace_add("write", self._apply_filter)

        # Audience selector — load last-used key, map back to display label
        _saved_key = _cfg.get("last_audience", "all")
        _saved_label = next(
            (lbl for lbl, key in self._AUDIENCE_OPTIONS.items() if key == _saved_key),
            "All (active + recently retired)",
        )
        self.audience_var = tk.StringVar(value=_saved_label)

        self._create_widgets()

        if _NEWSLETTER_AVAILABLE:
            self._scan_newsletter_files()
            self._refresh_recipients()
        else:
            self.toolbar_status_var.set(
                "Error: send_newsletter.py or the 'requests' package is not available."
            )
            self.preview_btn.config(state="disabled")

    # ── Layout ────────────────────────────────────────────────────────────────

    def _create_widgets(self):
        # Blue header bar
        hdr = tk.Frame(self, bg=self._BLUE, height=45)
        hdr.pack(fill="x")
        hdr.pack_propagate(False)
        tk.Label(hdr, text="Send Email Campaign", bg=self._BLUE, fg=self._WHITE,
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=15, pady=10)

        # Main content area
        content = ttk.Frame(self, padding=(12, 8, 12, 8))
        content.pack(fill="both", expand=True)

        # ── File picker ───────────────────────────────────────────────────────
        file_lf = ttk.LabelFrame(content, text="Campaign", padding=8)
        file_lf.pack(fill="x", pady=(0, 6))

        file_row = ttk.Frame(file_lf)
        file_row.pack(fill="x")

        self.file_var = tk.StringVar()
        self.file_combo = ttk.Combobox(file_row, textvariable=self.file_var,
                                        state="readonly", width=44)
        self.file_combo.pack(side="left", padx=(0, 6))
        self.file_combo.bind("<<ComboboxSelected>>", self._on_file_selected)

        self.browse_btn = ttk.Button(file_row, text="Browse...",
                                      command=self._browse_file)
        self.browse_btn.pack(side="left")

        subj_row = ttk.Frame(file_lf)
        subj_row.pack(fill="x", pady=(4, 0))
        ttk.Label(subj_row, text="Subject:", foreground="gray").pack(side="left")
        self.subject_label = ttk.Label(subj_row, text="—", foreground="gray",
                                        font=("Segoe UI", 9, "italic"))
        self.subject_label.pack(side="left", padx=(5, 0))

        aud_row = ttk.Frame(file_lf)
        aud_row.pack(fill="x", pady=(4, 0))
        ttk.Label(aud_row, text="Audience:").pack(side="left")
        self.audience_combo = ttk.Combobox(
            aud_row, textvariable=self.audience_var,
            values=list(self._AUDIENCE_OPTIONS.keys()),
            state="readonly", width=34,
        )
        self.audience_combo.pack(side="left", padx=(5, 0))
        # Trace fires after widget exists — wire after pack
        self.audience_var.trace_add("write", self._on_audience_changed)

        # ── Recipient counts ──────────────────────────────────────────────────
        recip_lf = ttk.LabelFrame(content, text="Recipients", padding=8)
        recip_lf.pack(fill="x", pady=(0, 6))

        counts_row = ttk.Frame(recip_lf)
        counts_row.pack(fill="x")

        ttk.Label(counts_row, text="Active:").pack(side="left")
        self.count_active_lbl = ttk.Label(counts_row, text="—",
                                           font=("Segoe UI", 9, "bold"))
        self.count_active_lbl.pack(side="left", padx=(3, 18))

        ttk.Label(counts_row, text="Recently retired:").pack(side="left")
        self.count_retired_lbl = ttk.Label(counts_row, text="—",
                                            font=("Segoe UI", 9, "bold"))
        self.count_retired_lbl.pack(side="left", padx=(3, 18))

        ttk.Label(counts_row, text="Total:").pack(side="left")
        self.count_total_lbl = ttk.Label(counts_row, text="—",
                                          font=("Segoe UI", 9, "bold"),
                                          foreground=self._BLUE)
        self.count_total_lbl.pack(side="left", padx=(3, 18))

        ttk.Label(counts_row, text="|", foreground="gray").pack(side="left", padx=(0, 18))
        ttk.Label(counts_row, text="Tagged:").pack(side="left")
        self.count_tagged_lbl = ttk.Label(counts_row, text="—",
                                           font=("Segoe UI", 9, "bold"),
                                           foreground=self._BLUE)
        self.count_tagged_lbl.pack(side="left", padx=(3, 0))

        self.count_multi_lbl = ttk.Label(recip_lf, text="", foreground="gray")
        self.count_multi_lbl.pack(anchor="w", pady=(3, 0))

        # ── Preview treeview (expands vertically) ─────────────────────────────
        prev_lf = ttk.LabelFrame(content, text="Preview", padding=8)
        prev_lf.pack(fill="both", expand=True, pady=(0, 6))

        # Filter / tag-control row
        filter_row = ttk.Frame(prev_lf)
        filter_row.pack(fill="x", pady=(0, 6))

        ttk.Label(filter_row, text="Filter:").pack(side="left")
        self.filter_var = tk.StringVar()
        self.filter_var.trace_add("write", self._apply_filter)
        ttk.Entry(filter_row, textvariable=self.filter_var,
                  width=28).pack(side="left", padx=(4, 4))
        ttk.Button(filter_row, text="Clear",
                   command=lambda: self.filter_var.set("")).pack(side="left", padx=(0, 10))
        ttk.Checkbutton(filter_row, text="Tagged Only",
                        variable=self.tagged_only_var).pack(side="left", padx=(0, 16))
        ttk.Button(filter_row, text="Tag All",
                   command=self._tag_all_visible).pack(side="left", padx=(0, 4))
        ttk.Button(filter_row, text="Untag All",
                   command=self._untag_all_visible).pack(side="left")

        cols = ("tagged", "status", "company", "name", "email")
        self.preview_tree = ttk.Treeview(prev_lf, columns=cols,
                                          show="headings", height=10)
        self.preview_tree.heading("tagged",  text="")
        self.preview_tree.heading("status",  text="Status")
        self.preview_tree.heading("company", text="Company")
        self.preview_tree.heading("name",    text="Name")
        self.preview_tree.heading("email",   text="Email Address(es)")

        self.preview_tree.column("tagged",  width=28,  minwidth=28,  stretch=False, anchor="center")
        self.preview_tree.column("status",  width=115, minwidth=90,  stretch=False)
        self.preview_tree.column("company", width=180, minwidth=80)
        self.preview_tree.column("name",    width=105, minwidth=70,  stretch=False)
        self.preview_tree.column("email",   width=230, minwidth=100)

        vsb = ttk.Scrollbar(prev_lf, orient="vertical",
                             command=self.preview_tree.yview)
        self.preview_tree.configure(yscrollcommand=vsb.set)
        self.preview_tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        # Subtle tint for recently-retired rows; click anywhere on a row to toggle tag
        self.preview_tree.tag_configure("retired", background="#F0F4E8")
        self.preview_tree.bind("<ButtonRelease-1>", self._on_tree_click)

        # ── Progress area (hidden until send starts) ──────────────────────────
        self._progress_frame = ttk.Frame(content)
        # Not packed here — inserted before _log_link when send begins.

        self._progress_bar = ttk.Progressbar(
            self._progress_frame, orient="horizontal", mode="determinate")
        self._progress_bar.pack(fill="x", pady=(0, 3))

        prog_detail = ttk.Frame(self._progress_frame)
        prog_detail.pack(fill="x")

        self._progress_status_var = tk.StringVar(value="")
        ttk.Label(prog_detail,
                  textvariable=self._progress_status_var).pack(side="left")

        self._sent_ok_var   = tk.IntVar(value=0)
        self._sent_fail_var = tk.IntVar(value=0)

        ttk.Label(prog_detail, text="Sent:",
                  foreground="green").pack(side="left", padx=(18, 2))
        ttk.Label(prog_detail, textvariable=self._sent_ok_var,
                  foreground="green",
                  font=("Segoe UI", 9, "bold")).pack(side="left")
        ttk.Label(prog_detail, text="Failed:",
                  foreground="red").pack(side="left", padx=(12, 2))
        ttk.Label(prog_detail, textvariable=self._sent_fail_var,
                  foreground="red",
                  font=("Segoe UI", 9, "bold")).pack(side="left")

        # ── Log link (empty until send completes) ─────────────────────────────
        self._log_link = tk.Label(content, text="", fg=self._BLUE,
                                   cursor="hand2", anchor="w",
                                   font=("Segoe UI", 9, "underline"))
        self._log_link.pack(fill="x")
        self._log_link.bind("<Button-1>", self._open_log)

        # ── Bottom toolbar ────────────────────────────────────────────────────
        toolbar = ttk.Frame(content)
        toolbar.pack(fill="x", pady=(4, 0))

        self.preview_btn = ttk.Button(toolbar, text="Preview",
                                       command=self._on_preview,
                                       style="Accent.TButton", width=10)
        self.preview_btn.pack(side="left", padx=(0, 6))

        self.send_btn = ttk.Button(toolbar, text="Send ▶",
                                    command=self._on_send, width=10)
        self.send_btn.config(state="disabled")
        self.send_btn.pack(side="left", padx=(0, 16))

        self.toolbar_status_var = tk.StringVar(
            value="Select a newsletter file, then click Preview.")
        ttk.Label(toolbar, textvariable=self.toolbar_status_var,
                  foreground="gray").pack(side="left")

        self.close_btn = ttk.Button(toolbar, text="Close",
                                     command=self._on_close, width=8)
        self.close_btn.pack(side="right")

        self.export_btn = ttk.Button(toolbar, text="Export…",
                                      command=self._on_export, width=9)
        self.export_btn.pack(side="right", padx=(0, 6))

    # ── File handling ─────────────────────────────────────────────────────────

    def _scan_newsletter_files(self):
        """Populate the file combo with newsletter-*.html files (newest first)."""
        pattern = os.path.join(str(APP_DIR), "newsletter-*.html")
        files   = sorted(_glob.glob(pattern), key=os.path.getmtime, reverse=True)
        names   = [os.path.basename(f) for f in files]
        self.file_combo["values"] = names
        if names:
            self.file_var.set(names[0])
            self._on_file_selected()
        else:
            self.toolbar_status_var.set(
                "No newsletter-*.html files found in the app folder. "
                "Use Browse to locate one."
            )

    def _get_selected_filepath(self) -> Optional[str]:
        """Return the absolute path to the selected newsletter file, or None."""
        fname = self.file_var.get().strip()
        if not fname:
            return None
        if os.path.isabs(fname) and os.path.isfile(fname):
            return fname
        fpath = os.path.join(str(APP_DIR), fname)
        return fpath if os.path.isfile(fpath) else None

    def _on_file_selected(self, event=None):
        """Update the subject preview and reset the send gate when the file changes."""
        fpath = self._get_selected_filepath()
        if fpath and _NEWSLETTER_AVAILABLE:
            try:
                with open(fpath, "r", encoding="utf-8") as f:
                    html = f.read()
                self.subject_label.config(
                    text=_newsletter_mod.get_subject_from_html(html))
            except Exception:
                self.subject_label.config(text="(could not read file)")
        elif not fpath:
            self.subject_label.config(text="(file not found)")

        if self._preview_done:
            self._preview_done = False
            self.send_btn.config(state="disabled")
            self.toolbar_status_var.set(
                "File changed — click Preview to confirm recipients, then Send.")

    def _browse_file(self):
        """Open a file picker for newsletter HTML files not in the app folder."""
        from tkinter import filedialog
        start_dir = self._last_folder or str(APP_DIR)
        path = filedialog.askopenfilename(
            title="Select Newsletter HTML",
            filetypes=[("HTML files", "*.html"), ("All files", "*.*")],
            initialdir=start_dir,
            parent=self,
        )
        if not path:
            return
        folder = os.path.dirname(path)
        self._last_folder = folder
        _save_config({**_load_config(), "newsletter_last_folder": folder})
        current = list(self.file_combo["values"])
        if path not in current:
            current.insert(0, path)
            self.file_combo["values"] = current
        self.file_var.set(path)
        self._on_file_selected()

    # ── Audience handling ─────────────────────────────────────────────────────

    def _current_audience_key(self) -> str:
        return self._AUDIENCE_OPTIONS.get(self.audience_var.get(), "all")

    def _on_audience_changed(self, *args):
        key = self._current_audience_key()
        _save_config({**_load_config(), "last_audience": key})
        self._refresh_recipients()
        if self._preview_done:
            self._preview_done = False
            self.send_btn.config(state="disabled")
            self.toolbar_status_var.set(
                "Audience changed — click Preview to confirm recipients, then Send.")

    # ── Recipient data ────────────────────────────────────────────────────────

    def _refresh_recipients(self):
        """Initial DB query on dialog open; updates count labels."""
        try:
            self._records = _newsletter_mod.get_send_list(
                _newsletter_mod.DB_PATH, self._current_audience_key())
            self._update_count_labels()
        except Exception as e:
            self.count_multi_lbl.config(text=f"Error loading recipients: {e}")

    def _update_count_labels(self):
        active  = sum(1 for r in self._records if r["send_status"] == "active")
        retired = sum(1 for r in self._records if r["send_status"] == "recently_retired")
        multi   = [r for r in self._records
                   if len(_newsletter_mod.extract_emails(
                       _newsletter_mod.get_primary_email_field(r))) > 1]
        extra   = sum(
            len(_newsletter_mod.extract_emails(
                _newsletter_mod.get_primary_email_field(r))) - 1
            for r in multi
        )
        self.count_active_lbl.config(text=str(active))
        self.count_retired_lbl.config(text=str(retired))
        self.count_total_lbl.config(text=str(len(self._records)))
        if multi:
            self.count_multi_lbl.config(
                text=(f"{len(multi)} multi-email account"
                      f"{'s' if len(multi) != 1 else ''} "
                      f"→ +{extra} extra send{'s' if extra != 1 else ''}")
            )
        else:
            self.count_multi_lbl.config(text="No multi-email accounts")
        self._update_tagged_label()

    # ── Preview ───────────────────────────────────────────────────────────────

    def _on_preview(self):
        if not _NEWSLETTER_AVAILABLE:
            messagebox.showerror(
                "Module Not Available",
                "send_newsletter.py or the 'requests' package is missing.",
                parent=self,
            )
            return

        self.toolbar_status_var.set("Loading recipients…")
        self.update_idletasks()

        try:
            self._records = _newsletter_mod.get_send_list(
                _newsletter_mod.DB_PATH, self._current_audience_key())
        except Exception as e:
            messagebox.showerror("Database Error",
                                  f"Could not load send list:\n{e}", parent=self)
            self.toolbar_status_var.set("Error — could not load recipients.")
            return

        # All rows tagged by default after every Preview run
        self._tagged = {str(i) for i in range(len(self._records))}

        self._update_count_labels()  # also calls _update_tagged_label

        # Clear the filter; the trace fires _apply_filter to build the tree
        self.filter_var.set("")

        self._preview_done = True
        self.send_btn.config(state="normal")
        self._update_status_bar()

    # ── Tag / filter helpers ──────────────────────────────────────────────────

    def _apply_filter(self, *args):
        """Re-render the preview treeview to only show rows matching the filter term."""
        if not self._records:
            return
        term = self.filter_var.get().strip().lower()
        for item in self.preview_tree.get_children():
            self.preview_tree.delete(item)
        for i, r in enumerate(self._records):
            iid       = str(i)
            company   = r.get("company_name") or ""
            contact   = r.get("contact_name") or ""
            emails    = _newsletter_mod.extract_emails(
                _newsletter_mod.get_primary_email_field(r))
            first     = _newsletter_mod.get_first_name(contact, company)
            email_str = ", ".join(emails) if emails else "(no valid email)"
            if term and term not in f"{company} {contact} {email_str}".lower():
                continue
            if self.tagged_only_var.get() and iid not in self._tagged:
                continue
            tag_char = "✓" if iid in self._tagged else ""
            tree_tag = "retired" if r["send_status"] == "recently_retired" else ""
            self.preview_tree.insert(
                "", "end", iid=iid,
                values=(tag_char, r["send_status"].replace("_", " "),
                        company, first, email_str),
                tags=(tree_tag,),
            )

    def _on_tree_click(self, event):
        """Toggle the tagged state of whichever row was clicked."""
        if self.preview_tree.identify_region(event.x, event.y) != "cell":
            return
        iid = self.preview_tree.identify_row(event.y)
        if iid:
            self._toggle_tag(iid)

    def _toggle_tag(self, iid: str):
        if iid in self._tagged:
            self._tagged.discard(iid)
            self.preview_tree.set(iid, "tagged", "")
        else:
            self._tagged.add(iid)
            self.preview_tree.set(iid, "tagged", "✓")
        self._update_tagged_label()

    def _tag_all_visible(self):
        """Tag every row currently visible in the treeview (respects active filter)."""
        for iid in self.preview_tree.get_children():
            self._tagged.add(iid)
            self.preview_tree.set(iid, "tagged", "✓")
        self._update_tagged_label()

    def _untag_all_visible(self):
        """Untag every row currently visible in the treeview (respects active filter)."""
        for iid in self.preview_tree.get_children():
            self._tagged.discard(iid)
            self.preview_tree.set(iid, "tagged", "")
        self._update_tagged_label()

    def _update_tagged_label(self):
        n = len(self._records)
        t = len(self._tagged)
        self.count_tagged_lbl.config(text=f"{t} of {n}" if n else "—")
        self._update_status_bar()

    def _update_status_bar(self):
        if not self._preview_done or not self._records:
            return
        n = len(self._records)
        t = len(self._tagged)
        s = "s" if n != 1 else ""
        if t == n:
            msg = f"Preview: {n} account{s}, all tagged.  Click Send ▶ when ready."
        elif t == 0:
            msg = f"Preview: {n} account{s}, 0 tagged.  Tag accounts before sending."
        else:
            msg = f"Preview: {n} account{s}, {t} tagged.  Click Send ▶ when ready."
        self.toolbar_status_var.set(msg)

    # ── Send ──────────────────────────────────────────────────────────────────

    def _on_send(self):
        api_key = os.environ.get("RESEND_API_KEY", "").strip()
        if not api_key:
            messagebox.showerror(
                "API Key Not Set",
                "The RESEND_API_KEY environment variable is not set.\n\n"
                "Set it before launching the CRM and try again.\n\n"
                "Example (Command Prompt):\n"
                "  set RESEND_API_KEY=re_xxxxxxxxxxxx",
                parent=self,
            )
            return

        fpath = self._get_selected_filepath()
        if not fpath:
            messagebox.showerror("No File Selected",
                                  "Please select a newsletter file.", parent=self)
            return

        try:
            with open(fpath, "r", encoding="utf-8") as f:
                html_template = f.read()
        except Exception as e:
            messagebox.showerror("File Error",
                                  f"Could not read newsletter file:\n{e}", parent=self)
            return

        tagged_records = [self._records[int(iid)]
                          for iid in sorted(self._tagged, key=int)]
        if not tagged_records:
            messagebox.showwarning(
                "Nothing Tagged",
                "No accounts are tagged.\n\n"
                "Use Tag All or click individual rows to tag accounts before sending.",
                parent=self,
            )
            return

        subject = _newsletter_mod.get_subject_from_html(html_template)
        n       = len(tagged_records)
        active  = sum(1 for r in tagged_records if r["send_status"] == "active")
        retired = sum(1 for r in tagged_records if r["send_status"] == "recently_retired")

        confirmed = messagebox.askyesno(
            "Confirm Send",
            f"You are about to send to {n} tagged account{'s' if n != 1 else ''} "
            f"({active} active, {retired} recently retired).\n\n"
            f"Subject: {subject}\n\n"
            "This cannot be undone. Continue?",
            icon="warning",
            parent=self,
        )
        if not confirmed:
            return

        # Reset counters
        self._sent_ok = self._sent_fail = self._sent_skipped = 0
        self._sent_ok_var.set(0)
        self._sent_fail_var.set(0)
        self._progress_bar["maximum"] = n
        self._progress_bar["value"]   = 0
        self._progress_status_var.set("Starting…")

        # Show progress frame above the log link
        self._progress_frame.pack(fill="x", pady=(0, 4), before=self._log_link)
        self._log_link.config(text="")

        # Lock all controls during send
        self._sending = True
        self.preview_btn.config(state="disabled")
        self.send_btn.config(state="disabled")
        self.close_btn.config(state="disabled")
        self.file_combo.config(state="disabled")
        self.browse_btn.config(state="disabled")
        self.audience_combo.config(state="disabled")
        self.toolbar_status_var.set("Sending…")

        threading.Thread(
            target=self._send_worker,
            args=(tagged_records, html_template, subject, api_key, fpath),
            daemon=True,
        ).start()

    def _send_worker(self, records, html_template, subject, api_key, filepath):
        """Background thread: iterate records, call Resend, post UI updates via after()."""
        from send_newsletter import (
            extract_emails, get_primary_email_field, get_first_name,
            build_unsubscribe_link, merge_template, send_email,
            open_log_file, log_row,
        )

        log_file, log_writer, log_path = open_log_file(filepath)
        n = len(records)

        for i, record in enumerate(records, 1):
            email_field = get_primary_email_field(record)
            emails      = extract_emails(email_field)
            first_name  = get_first_name(
                record.get("contact_name"), record.get("company_name"))

            if not emails:
                log_row(log_writer, record,
                        email_field or "(empty)", "skipped", "no valid email address")
                self.after(0, self._tick_skipped)
            else:
                company_name = record.get("company_name") or ""
                for email in emails:
                    unsub     = build_unsubscribe_link(email)
                    merged    = merge_template(html_template, first_name, unsub, company_name)
                    ok, error = send_email(email, subject, merged, api_key)
                    log_row(log_writer, record, email,
                            "sent" if ok else "failed", error)
                    self.after(0, self._tick_result, ok)

            self.after(0, self._tick_progress, i, n)

        log_file.close()
        self.after(0, self._on_send_complete, log_path)

    # -- Thread callbacks (always invoked on the main thread via after()) ------

    def _tick_result(self, success: bool):
        if success:
            self._sent_ok += 1
            self._sent_ok_var.set(self._sent_ok)
        else:
            self._sent_fail += 1
            self._sent_fail_var.set(self._sent_fail)

    def _tick_skipped(self):
        self._sent_skipped += 1

    def _tick_progress(self, current: int, total: int):
        self._progress_bar["value"] = current
        self._progress_status_var.set(f"Account {current} of {total}…")

    def _on_send_complete(self, log_path: str):
        self._sending      = False
        self._log_path     = log_path
        self._preview_done = False   # Re-preview required before next send

        self.preview_btn.config(state="normal")
        self.send_btn.config(state="disabled")   # gated until preview re-run
        self.close_btn.config(state="normal")
        self.file_combo.config(state="readonly")
        self.browse_btn.config(state="normal")
        self.audience_combo.config(state="readonly")

        self._progress_status_var.set("Complete")

        summary = f"Done — Sent: {self._sent_ok},  Failed: {self._sent_fail}"
        if self._sent_skipped:
            summary += f",  Skipped: {self._sent_skipped}"
        if self._sent_fail:
            summary += "  — filter result=failed in log for retry list"
        self.toolbar_status_var.set(summary)

        log_name = os.path.basename(log_path)
        self._log_link.config(text=f"Open send log: {log_name}")

        msg = (
            f"Send complete.\n\n"
            f"Sent OK:  {self._sent_ok}\n"
            f"Failed:    {self._sent_fail}"
        )
        if self._sent_skipped:
            msg += f"\nSkipped:  {self._sent_skipped}  (no valid email address)"
        msg += f"\n\nLog saved to:\n{log_path}"
        messagebox.showinfo("Send Complete", msg, parent=self)

    # ── Export ────────────────────────────────────────────────────────────────

    def _on_export(self):
        if not self._records:
            messagebox.showinfo(
                "No Data",
                "Click Preview first to load recipients.",
                parent=self,
            )
            return

        import csv
        from tkinter import filedialog

        visible = self.preview_tree.get_children()
        if not visible:
            messagebox.showinfo(
                "Nothing to Export",
                "No rows are visible. Adjust the filter or run Preview.",
                parent=self,
            )
            return

        default_name = f"campaign_export_{datetime.now().strftime('%Y-%m-%d')}.csv"
        reports_dir  = os.path.join(str(APP_DIR), "Reports")
        os.makedirs(reports_dir, exist_ok=True)
        start_dir    = _load_config().get("export_last_folder", "") or reports_dir
        path = filedialog.asksaveasfilename(
            title="Export Visible Rows",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=default_name,
            initialdir=start_dir,
            parent=self,
        )
        if not path:
            return

        rows_written = 0
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["Tagged", "Status", "Company", "Name", "Email"])
            for iid in visible:
                r        = self._records[int(iid)]
                company  = r.get("company_name") or ""
                contact  = r.get("contact_name") or ""
                emails   = _newsletter_mod.extract_emails(
                    _newsletter_mod.get_primary_email_field(r))
                name     = _newsletter_mod.get_first_name(contact, company)
                email_str = ", ".join(emails) if emails else "(no valid email)"
                writer.writerow([
                    "Yes" if iid in self._tagged else "No",
                    r["send_status"].replace("_", " "),
                    company,
                    name,
                    email_str,
                ])
                rows_written += 1

        _save_config({**_load_config(), "export_last_folder": os.path.dirname(path)})

        fname = os.path.basename(path)
        s     = "s" if rows_written != 1 else ""

        dlg = tk.Toplevel(self)
        dlg.title("Export Complete")
        dlg.resizable(False, False)
        dlg.transient(self)
        dlg.grab_set()

        ttk.Label(dlg, text=f"Exported {rows_written} row{s} to:",
                  padding=(16, 14, 16, 2)).pack(anchor="w")
        ttk.Label(dlg, text=fname, font=("Segoe UI", 9, "bold"),
                  padding=(16, 0, 16, 14)).pack(anchor="w")

        btn_row = ttk.Frame(dlg, padding=(12, 0, 12, 12))
        btn_row.pack(fill="x")
        ttk.Button(btn_row, text="Open File",
                   style="Accent.TButton",
                   command=lambda: (os.startfile(path), dlg.destroy())
                   ).pack(side="left", padx=(0, 6))
        ttk.Button(btn_row, text="Close",
                   command=dlg.destroy).pack(side="left")

        dlg.update_idletasks()
        x = self.winfo_x() + (self.winfo_width()  - dlg.winfo_width())  // 2
        y = self.winfo_y() + (self.winfo_height() - dlg.winfo_height()) // 2
        dlg.geometry(f"+{x}+{max(0, y)}")
        dlg.wait_window()

    # ── Log link ──────────────────────────────────────────────────────────────

    def _open_log(self, event=None):
        if self._log_path and os.path.isfile(self._log_path):
            os.startfile(self._log_path)

    # ── Window close ──────────────────────────────────────────────────────────

    def _on_close(self):
        if self._sending:
            if not messagebox.askyesno(
                "Send in Progress",
                "A newsletter send is currently running.\n\n"
                "Closing now will interrupt it — accounts not yet processed "
                "will not receive the email.\n\n"
                "Close anyway?",
                icon="warning",
                parent=self,
            ):
                return
        self.destroy()
