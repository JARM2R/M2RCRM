"""
M2R CRM Contact Export Tool
============================
Exports CRM contacts to Excel with date filtering based on Paid Through date.

Columns exported:
- First Name, Last Name, Company Name, Email Address, Phone Number
- Street Address 1, Street Address 2, City, State, Postal Code
- Reference ID, Birthday, Email Subscription Status

Usage:
    python crm_contact_export.py

Requirements:
    pip install pyodbc openpyxl tkcalendar
"""

import sys
import os
from datetime import datetime, date
from pathlib import Path
from typing import List, Dict, Optional, Tuple

try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
except ImportError:
    print("ERROR: tkinter not available")
    sys.exit(1)

try:
    import pyodbc
except ImportError:
    print("ERROR: pyodbc not installed. Run: pip install pyodbc")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

try:
    from tkcalendar import DateEntry
    CALENDAR_AVAILABLE = True
except ImportError:
    CALENDAR_AVAILABLE = False
    print("Note: tkcalendar not installed. Using manual date entry. Run: pip install tkcalendar")


from m2r_crm_paths import APP_DIR

# Default database path (writable directory — exe folder when frozen, source folder in dev)
DEFAULT_DB_PATH = os.path.join(str(APP_DIR), "M2R 2021 CRM.accdb")

# Excel column headers (exact order as specified)
EXPORT_COLUMNS = [
    "First Name",
    "Last Name",
    "Company Name",
    "Email Address",
    "Phone Number",
    "Street Address 1",
    "Street Address 2",
    "City",
    "State",
    "Postal Code",
    "Reference ID",
    "Birthday",
    "Email Subscription Status"
]

# Possible field name mappings from Access database to export columns
# Format: Export Column -> List of possible Access field names
FIELD_MAPPINGS = {
    "First Name": ["FirstName", "First Name", "ContactFirstName", "First_Name", "FName"],
    "Last Name": ["LastName", "Last Name", "ContactLastName", "Last_Name", "LName", "Surname"],
    "Company Name": ["CompanyName", "Company Name", "Company", "BusinessName", "Organization"],
    "Email Address": ["Email", "EmailAddress", "Email Address", "E-mail", "ContactEmail"],
    "Phone Number": ["Phone", "PhoneNumber", "Phone Number", "Telephone", "ContactPhone", "BusinessPhone"],
    "Street Address 1": ["Address1", "Address 1", "Street Address 1", "StreetAddress1", "AddressLine1", "Address"],
    "Street Address 2": ["Address2", "Address 2", "Street Address 2", "StreetAddress2", "AddressLine2"],
    "City": ["City", "ContactCity"],
    "State": ["State", "StateProvince", "Province", "ContactState"],
    "Postal Code": ["PostalCode", "Postal Code", "ZipCode", "Zip Code", "ZIP", "Zip"],
    "Reference ID": ["ReferenceID", "Reference ID", "RefID", "ID", "ContactID", "CustomerID", "AccountID"],
    "Birthday": ["Birthday", "BirthDate", "Birth Date", "DateOfBirth", "DOB"],
    "Email Subscription Status": ["EmailSubscription", "Email Subscription Status", "SubscriptionStatus",
                                   "EmailOptIn", "Newsletter", "EmailStatus", "Subscribed"]
}

# Possible "Paid Through" date field names
PAID_THROUGH_FIELDS = [
    "PaidThrough", "Paid Through", "PaidThroughDate", "Paid Through Date",
    "PaidThru", "PaidToDate", "SubscriptionEnd", "ExpirationDate",
    "Expiration", "RenewalDate", "ServiceEndDate", "AccountExpires"
]


def get_access_connection_string(db_path: str) -> str:
    """Get ODBC connection string for Access database."""
    abs_path = os.path.abspath(db_path)

    if not os.path.exists(abs_path):
        raise FileNotFoundError(f"Database not found: {abs_path}")

    drivers = [d for d in pyodbc.drivers() if 'Access' in d]

    for driver in drivers:
        try:
            conn_str = f'DRIVER={{{driver}}};DBQ={abs_path};'
            conn = pyodbc.connect(conn_str, timeout=5)
            conn.close()
            return conn_str
        except pyodbc.Error:
            continue

    raise RuntimeError(
        "No Access ODBC driver found. Install Microsoft Access Database Engine:\n"
        "https://www.microsoft.com/en-us/download/details.aspx?id=54920"
    )


def get_table_names(conn_str: str) -> List[str]:
    """Get list of user tables in the database."""
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    tables = []
    for row in cursor.tables(tableType='TABLE'):
        table_name = row.table_name
        # Skip system tables
        if not table_name.startswith('MSys') and not table_name.startswith('~'):
            tables.append(table_name)

    cursor.close()
    conn.close()
    return sorted(tables)


def get_table_columns(conn_str: str, table_name: str) -> List[str]:
    """Get list of column names for a table."""
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    columns = []
    for row in cursor.columns(table=table_name):
        columns.append(row.column_name)

    cursor.close()
    conn.close()
    return columns


def find_matching_column(db_columns: List[str], possible_names: List[str]) -> Optional[str]:
    """Find matching column name from database columns."""
    db_columns_lower = {col.lower(): col for col in db_columns}

    for name in possible_names:
        name_lower = name.lower().replace(" ", "").replace("_", "")
        for db_col_lower, db_col in db_columns_lower.items():
            if db_col_lower.replace(" ", "").replace("_", "") == name_lower:
                return db_col
    return None


def detect_paid_through_column(db_columns: List[str]) -> Optional[str]:
    """Detect the Paid Through date column."""
    return find_matching_column(db_columns, PAID_THROUGH_FIELDS)


def build_column_mapping(db_columns: List[str]) -> Dict[str, Optional[str]]:
    """Build mapping from export columns to database columns."""
    mapping = {}
    for export_col, possible_names in FIELD_MAPPINGS.items():
        mapping[export_col] = find_matching_column(db_columns, possible_names)
    return mapping


def query_contacts(conn_str: str, table_name: str, column_mapping: Dict[str, Optional[str]],
                   paid_through_col: Optional[str], start_date: Optional[date],
                   end_date: Optional[date]) -> List[Dict]:
    """Query contacts from database with optional date filtering."""
    conn = pyodbc.connect(conn_str)
    cursor = conn.cursor()

    # Build SELECT clause
    select_cols = []
    col_aliases = []
    for export_col in EXPORT_COLUMNS:
        db_col = column_mapping.get(export_col)
        if db_col:
            select_cols.append(f"[{db_col}]")
            col_aliases.append(export_col)
        else:
            select_cols.append("NULL")
            col_aliases.append(export_col)

    # Build query
    sql = f"SELECT {', '.join(select_cols)} FROM [{table_name}]"
    params = []

    # Add date filter if Paid Through column exists and dates are provided
    if paid_through_col and (start_date or end_date):
        conditions = []
        if start_date:
            conditions.append(f"[{paid_through_col}] >= ?")
            params.append(start_date)
        if end_date:
            conditions.append(f"[{paid_through_col}] <= ?")
            params.append(end_date)

        if conditions:
            sql += " WHERE " + " AND ".join(conditions)

    sql += f" ORDER BY [{column_mapping.get('Last Name') or column_mapping.get('Company Name') or 'ID'}]"

    try:
        cursor.execute(sql, params)
        rows = cursor.fetchall()
    except pyodbc.Error as e:
        cursor.close()
        conn.close()
        raise RuntimeError(f"Query error: {e}")

    # Convert to list of dicts
    results = []
    for row in rows:
        record = {}
        for i, col_name in enumerate(col_aliases):
            value = row[i]
            # Format dates and handle None
            if value is None:
                record[col_name] = ""
            elif isinstance(value, datetime):
                record[col_name] = value.strftime("%m/%d/%Y")
            elif isinstance(value, date):
                record[col_name] = value.strftime("%m/%d/%Y")
            elif isinstance(value, bool):
                record[col_name] = "Yes" if value else "No"
            else:
                record[col_name] = str(value).strip()
        results.append(record)

    cursor.close()
    conn.close()
    return results


def export_to_excel(contacts: List[Dict], output_path: str,
                    start_date: Optional[date] = None, end_date: Optional[date] = None) -> int:
    """Export contacts to Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, contact in enumerate(contacts, 2):
        for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
            value = contact.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
        max_length = len(header)
        for row_idx in range(2, len(contacts) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))

        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add metadata sheet
    meta_ws = wb.create_sheet("Export Info")
    meta_ws.cell(row=1, column=1, value="Export Date:").font = Font(bold=True)
    meta_ws.cell(row=1, column=2, value=datetime.now().strftime("%m/%d/%Y %I:%M %p"))
    meta_ws.cell(row=2, column=1, value="Total Records:").font = Font(bold=True)
    meta_ws.cell(row=2, column=2, value=len(contacts))

    if start_date or end_date:
        meta_ws.cell(row=3, column=1, value="Paid Through Filter:").font = Font(bold=True)
        date_range = []
        if start_date:
            date_range.append(f"From: {start_date.strftime('%m/%d/%Y')}")
        if end_date:
            date_range.append(f"To: {end_date.strftime('%m/%d/%Y')}")
        meta_ws.cell(row=3, column=2, value=" | ".join(date_range))

    meta_ws.column_dimensions['A'].width = 20
    meta_ws.column_dimensions['B'].width = 30

    wb.save(output_path)
    return len(contacts)


class ContactExportApp:
    """Main application window."""

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("M2R CRM Contact Export")
        self.root.geometry("600x500")
        self.root.resizable(True, True)

        self.conn_str = None
        self.table_name = None
        self.column_mapping = {}
        self.paid_through_col = None
        self.db_columns = []

        self._create_ui()
        self._load_default_database()

    def _create_ui(self):
        """Create the user interface."""
        main_frame = ttk.Frame(self.root, padding=15)
        main_frame.pack(fill="both", expand=True)

        # Title
        title_label = ttk.Label(main_frame, text="CRM Contact Export", font=("Segoe UI", 16, "bold"))
        title_label.pack(anchor="w", pady=(0, 10))

        # Database section
        db_frame = ttk.LabelFrame(main_frame, text="Database Connection", padding=10)
        db_frame.pack(fill="x", pady=(0, 10))

        db_path_frame = ttk.Frame(db_frame)
        db_path_frame.pack(fill="x")

        ttk.Label(db_path_frame, text="Database:").pack(side="left")
        self.db_path_var = tk.StringVar(value=DEFAULT_DB_PATH)
        self.db_path_entry = ttk.Entry(db_path_frame, textvariable=self.db_path_var, width=50)
        self.db_path_entry.pack(side="left", padx=(5, 5), fill="x", expand=True)

        browse_btn = ttk.Button(db_path_frame, text="Browse...", command=self._browse_database)
        browse_btn.pack(side="left")

        # Table selection
        table_frame = ttk.Frame(db_frame)
        table_frame.pack(fill="x", pady=(10, 0))

        ttk.Label(table_frame, text="Table:").pack(side="left")
        self.table_var = tk.StringVar()
        self.table_combo = ttk.Combobox(table_frame, textvariable=self.table_var, state="readonly", width=40)
        self.table_combo.pack(side="left", padx=(5, 5))
        self.table_combo.bind("<<ComboboxSelected>>", self._on_table_selected)

        refresh_btn = ttk.Button(table_frame, text="Refresh", command=self._refresh_tables)
        refresh_btn.pack(side="left")

        # Date filter section
        filter_frame = ttk.LabelFrame(main_frame, text="Paid Through Date Filter", padding=10)
        filter_frame.pack(fill="x", pady=(0, 10))

        date_row = ttk.Frame(filter_frame)
        date_row.pack(fill="x")

        # Start date
        ttk.Label(date_row, text="From:").pack(side="left")

        if CALENDAR_AVAILABLE:
            self.start_date_entry = DateEntry(date_row, width=12, date_pattern='mm/dd/yyyy')
            self.start_date_entry.pack(side="left", padx=(5, 15))
            self.start_date_entry.delete(0, "end")  # Clear default date
        else:
            self.start_date_var = tk.StringVar()
            self.start_date_entry = ttk.Entry(date_row, textvariable=self.start_date_var, width=12)
            self.start_date_entry.pack(side="left", padx=(5, 5))
            ttk.Label(date_row, text="(MM/DD/YYYY)").pack(side="left", padx=(0, 10))

        # End date
        ttk.Label(date_row, text="To:").pack(side="left")

        if CALENDAR_AVAILABLE:
            self.end_date_entry = DateEntry(date_row, width=12, date_pattern='mm/dd/yyyy')
            self.end_date_entry.pack(side="left", padx=(5, 15))
            self.end_date_entry.delete(0, "end")  # Clear default date
        else:
            self.end_date_var = tk.StringVar()
            self.end_date_entry = ttk.Entry(date_row, textvariable=self.end_date_var, width=12)
            self.end_date_entry.pack(side="left", padx=(5, 5))
            ttk.Label(date_row, text="(MM/DD/YYYY)").pack(side="left", padx=(0, 10))

        clear_btn = ttk.Button(date_row, text="Clear Dates", command=self._clear_dates)
        clear_btn.pack(side="left", padx=(10, 0))

        # Paid Through column indicator
        self.paid_through_label = ttk.Label(filter_frame, text="Paid Through column: Not detected", foreground="gray")
        self.paid_through_label.pack(anchor="w", pady=(5, 0))

        # Column mapping info
        mapping_frame = ttk.LabelFrame(main_frame, text="Column Mapping Status", padding=10)
        mapping_frame.pack(fill="both", expand=True, pady=(0, 10))

        # Create scrollable text for mapping info
        mapping_scroll = ttk.Scrollbar(mapping_frame)
        mapping_scroll.pack(side="right", fill="y")

        self.mapping_text = tk.Text(mapping_frame, height=10, width=60, state="disabled",
                                     yscrollcommand=mapping_scroll.set, font=("Consolas", 9))
        self.mapping_text.pack(fill="both", expand=True)
        mapping_scroll.config(command=self.mapping_text.yview)

        # Export button section
        export_frame = ttk.Frame(main_frame)
        export_frame.pack(fill="x", pady=(0, 5))

        self.export_btn = ttk.Button(export_frame, text="Export to Excel", command=self._export,
                                      style="Accent.TButton")
        self.export_btn.pack(side="left")

        self.status_var = tk.StringVar(value="Ready")
        status_label = ttk.Label(export_frame, textvariable=self.status_var)
        status_label.pack(side="left", padx=(15, 0))

    def _browse_database(self):
        """Browse for database file."""
        path = filedialog.askopenfilename(
            title="Select Access Database",
            filetypes=[("Access Database", "*.accdb *.mdb"), ("All Files", "*.*")],
            initialdir=os.path.dirname(self.db_path_var.get())
        )
        if path:
            self.db_path_var.set(path)
            self._refresh_tables()

    def _load_default_database(self):
        """Load the default database on startup."""
        if os.path.exists(DEFAULT_DB_PATH):
            self._refresh_tables()

    def _refresh_tables(self):
        """Refresh the list of tables from the database."""
        db_path = self.db_path_var.get()

        if not os.path.exists(db_path):
            messagebox.showerror("Error", f"Database file not found:\n{db_path}")
            return

        try:
            self.conn_str = get_access_connection_string(db_path)
            tables = get_table_names(self.conn_str)

            self.table_combo['values'] = tables

            if tables:
                # Try to auto-select a likely contacts table
                likely_names = ['contacts', 'customers', 'clients', 'accounts', 'members']
                selected = None
                for name in likely_names:
                    for table in tables:
                        if name in table.lower():
                            selected = table
                            break
                    if selected:
                        break

                if selected:
                    self.table_combo.set(selected)
                else:
                    self.table_combo.set(tables[0])

                self._on_table_selected(None)

            self.status_var.set(f"Connected - {len(tables)} tables found")

        except Exception as e:
            messagebox.showerror("Database Error", str(e))
            self.status_var.set("Connection failed")

    def _on_table_selected(self, event):
        """Handle table selection change."""
        self.table_name = self.table_var.get()

        if not self.table_name or not self.conn_str:
            return

        try:
            self.db_columns = get_table_columns(self.conn_str, self.table_name)
            self.column_mapping = build_column_mapping(self.db_columns)
            self.paid_through_col = detect_paid_through_column(self.db_columns)

            self._update_mapping_display()

            if self.paid_through_col:
                self.paid_through_label.config(
                    text=f"Paid Through column: {self.paid_through_col}",
                    foreground="green"
                )
            else:
                self.paid_through_label.config(
                    text="Paid Through column: Not detected (date filter disabled)",
                    foreground="orange"
                )

        except Exception as e:
            messagebox.showerror("Error", f"Failed to read table structure:\n{e}")

    def _update_mapping_display(self):
        """Update the column mapping display."""
        self.mapping_text.config(state="normal")
        self.mapping_text.delete("1.0", "end")

        mapped_count = 0
        for export_col in EXPORT_COLUMNS:
            db_col = self.column_mapping.get(export_col)
            if db_col:
                status = f"[OK]  {export_col}  ->  {db_col}"
                mapped_count += 1
            else:
                status = f"[--]  {export_col}  ->  (not found - will be blank)"
            self.mapping_text.insert("end", status + "\n")

        self.mapping_text.insert("end", f"\n{mapped_count}/{len(EXPORT_COLUMNS)} columns mapped")
        self.mapping_text.config(state="disabled")

    def _clear_dates(self):
        """Clear the date filter entries."""
        if CALENDAR_AVAILABLE:
            self.start_date_entry.delete(0, "end")
            self.end_date_entry.delete(0, "end")
        else:
            self.start_date_var.set("")
            self.end_date_var.set("")

    def _get_date(self, entry) -> Optional[date]:
        """Get date from entry widget."""
        if CALENDAR_AVAILABLE:
            try:
                date_str = entry.get()
                if not date_str:
                    return None
                return entry.get_date()
            except:
                return None
        else:
            date_str = entry.get().strip()
            if not date_str:
                return None

            for fmt in ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d']:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return None

    def _export(self):
        """Export contacts to Excel."""
        if not self.conn_str or not self.table_name:
            messagebox.showerror("Error", "Please select a database and table first.")
            return

        # Get date filters
        start_date = self._get_date(self.start_date_entry)
        end_date = self._get_date(self.end_date_entry)

        # Validate dates
        if start_date and end_date and start_date > end_date:
            messagebox.showerror("Error", "Start date cannot be after end date.")
            return

        # If dates provided but no Paid Through column, warn user
        if (start_date or end_date) and not self.paid_through_col:
            result = messagebox.askyesno(
                "Warning",
                "No 'Paid Through' column was detected in this table.\n"
                "Date filtering will be disabled.\n\n"
                "Continue without date filter?"
            )
            if not result:
                return
            start_date = None
            end_date = None

        # Choose output file
        default_name = f"Contact_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=default_name,
            initialdir=os.path.dirname(self.db_path_var.get())
        )

        if not output_path:
            return

        try:
            self.status_var.set("Querying database...")
            self.root.update_idletasks()

            contacts = query_contacts(
                self.conn_str,
                self.table_name,
                self.column_mapping,
                self.paid_through_col,
                start_date,
                end_date
            )

            if not contacts:
                messagebox.showwarning("No Data", "No contacts found matching the criteria.")
                self.status_var.set("Ready")
                return

            self.status_var.set(f"Exporting {len(contacts)} contacts...")
            self.root.update_idletasks()

            count = export_to_excel(contacts, output_path, start_date, end_date)

            self.status_var.set(f"Exported {count} contacts")

            messagebox.showinfo(
                "Export Complete",
                f"Successfully exported {count} contacts to:\n{output_path}"
            )

            # Offer to open the file
            if messagebox.askyesno("Open File?", "Would you like to open the exported file?"):
                os.startfile(output_path)

        except Exception as e:
            messagebox.showerror("Export Error", str(e))
            self.status_var.set("Export failed")


def main():
    """Main entry point."""
    root = tk.Tk()

    # Set theme
    style = ttk.Style()
    if 'clam' in style.theme_names():
        style.theme_use('clam')

    app = ContactExportApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
