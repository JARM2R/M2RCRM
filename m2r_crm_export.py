"""
M2R CRM Export Module
=====================
Excel export functionality for the M2R CRM application.
"""

import os
from datetime import datetime, date
from typing import List, Dict, Optional

import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from tkcalendar import DateEntry
    CALENDAR_AVAILABLE = True
except ImportError:
    CALENDAR_AVAILABLE = False

import m2r_crm_database as db


# Excel column headers for export
EXPORT_COLUMNS = [
    "First Name",
    "Last Name",
    "Company Name",
    "AP Email",
    "User Email",
    "IT Email",
    "Phone Number",
    "Street Address 1",
    "Street Address 2",
    "City",
    "State",
    "Postal Code",
    "Reference ID",
    "Birthday",
    "Email Subscription Status",
    "Account Number",
    "Company ID",
    "Equifax ID",
    "Experian ID",
    "TransUnion ID",
    "Paid Through Date",
    "Support Amt"
]


def split_name(full_name: str) -> tuple:
    """Split a full name into first and last name."""
    if not full_name:
        return ("", "")

    parts = full_name.strip().split(None, 1)
    if len(parts) == 1:
        return (parts[0], "")
    return (parts[0], parts[1])


def format_contact_for_export(customer: Dict) -> Dict:
    """Format a customer record for Excel export."""
    first_name, last_name = split_name(customer.get('contact_name', ''))

    # Determine subscription status
    status = customer.get('subscription_status', 'NO')
    email_status = "Subscribed" if status == 'YES' else "Unsubscribed"

    # Format paid through date
    paid_through = customer.get('paid_through_date', '')
    if paid_through:
        try:
            if isinstance(paid_through, str):
                paid_through_formatted = paid_through
            else:
                paid_through_formatted = paid_through.strftime('%m/%d/%Y')
        except:
            paid_through_formatted = str(paid_through)
    else:
        paid_through_formatted = ''

    # Format support amount
    support_amt = customer.get('support_amount', '')
    if support_amt is not None and support_amt != '':
        try:
            support_amt_formatted = f"${float(support_amt):.2f}"
        except:
            support_amt_formatted = str(support_amt)
    else:
        support_amt_formatted = ''

    return {
        "First Name": first_name,
        "Last Name": last_name,
        "Company Name": customer.get('company_name', ''),
        "AP Email": customer.get('email_ap', '') or customer.get('email', ''),
        "User Email": customer.get('email_user', '') or '',
        "IT Email": customer.get('email_it', '') or '',
        "Phone Number": customer.get('phone', ''),
        "Street Address 1": customer.get('address_street', ''),
        "Street Address 2": "",  # We don't have a separate field for this
        "City": customer.get('address_city', ''),
        "State": customer.get('address_state', ''),
        "Postal Code": customer.get('address_zip', ''),
        "Reference ID": customer.get('account_number', ''),
        "Birthday": "",  # We don't track birthdays
        "Email Subscription Status": email_status,
        "Account Number": customer.get('account_number', ''),
        "Company ID": customer.get('company_id', '') or '',
        "Equifax ID": customer.get('bureau_equifax', '') or '',
        "Experian ID": customer.get('bureau_experian', '') or '',
        "TransUnion ID": customer.get('bureau_transunion', '') or '',
        "Paid Through Date": paid_through_formatted,
        "Support Amt": support_amt_formatted
    }


def export_to_excel(
    contacts: List[Dict],
    output_path: str,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None
) -> int:
    """
    Export contacts to Excel file.

    Args:
        contacts: List of contact dictionaries
        output_path: Path for the output Excel file
        start_date: Optional start date filter (for metadata)
        end_date: Optional end date filter (for metadata)

    Returns:
        Number of contacts exported
    """
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl is required for Excel export. Install with: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Contacts"

    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, contact in enumerate(contacts, 2):
        formatted = format_contact_for_export(contact)
        for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
            value = formatted.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col_idx, header in enumerate(EXPORT_COLUMNS, 1):
        max_length = len(header)
        for row_idx in range(2, len(contacts) + 2):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add metadata sheet
    meta_ws = wb.create_sheet("Export Info")
    meta_ws.cell(row=1, column=1, value="Export Date:").font = Font(bold=True)
    meta_ws.cell(row=1, column=2, value=datetime.now().strftime("%m/%d/%Y %I:%M %p"))
    meta_ws.cell(row=2, column=1, value="Total Records:").font = Font(bold=True)
    meta_ws.cell(row=2, column=2, value=len(contacts))

    if start_date or end_date:
        meta_ws.cell(row=3, column=1, value="Paid Through Filter:").font = Font(bold=True)
        date_range = []
        if start_date:
            date_range.append(f"From: {start_date.strftime('%m/%d/%Y')}")
        if end_date:
            date_range.append(f"To: {end_date.strftime('%m/%d/%Y')}")
        meta_ws.cell(row=3, column=2, value=" | ".join(date_range))

    meta_ws.column_dimensions['A'].width = 20
    meta_ws.column_dimensions['B'].width = 30

    wb.save(output_path)
    return len(contacts)


class ExportDialog(tk.Toplevel):
    """Dialog for exporting contacts to Excel."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Export Contacts")
        self.geometry("450x300")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()

        # Center on parent
        self.update_idletasks()
        x = parent.winfo_x() + (parent.winfo_width() - 450) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 300) // 2
        self.geometry(f"+{x}+{y}")

        self._create_widgets()

    def _create_widgets(self):
        main_frame = ttk.Frame(self, padding=20)
        main_frame.pack(fill="both", expand=True)

        # Title
        ttk.Label(main_frame, text="Export Contacts to Excel",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", pady=(0, 15))

        # Date filter frame
        filter_frame = ttk.LabelFrame(main_frame, text="Paid Through Date Filter", padding=10)
        filter_frame.pack(fill="x", pady=(0, 15))

        date_row = ttk.Frame(filter_frame)
        date_row.pack(fill="x")

        # Start date
        ttk.Label(date_row, text="From:").pack(side="left")
        if CALENDAR_AVAILABLE:
            self.start_date_entry = DateEntry(date_row, width=12, date_pattern='mm/dd/yyyy')
            self.start_date_entry.pack(side="left", padx=(5, 15))
            self.start_date_entry.delete(0, "end")
        else:
            self.start_date_var = tk.StringVar()
            self.start_date_entry = ttk.Entry(date_row, textvariable=self.start_date_var, width=12)
            self.start_date_entry.pack(side="left", padx=(5, 15))

        # End date
        ttk.Label(date_row, text="To:").pack(side="left")
        if CALENDAR_AVAILABLE:
            self.end_date_entry = DateEntry(date_row, width=12, date_pattern='mm/dd/yyyy')
            self.end_date_entry.pack(side="left", padx=(5, 15))
            self.end_date_entry.delete(0, "end")
        else:
            self.end_date_var = tk.StringVar()
            self.end_date_entry = ttk.Entry(date_row, textvariable=self.end_date_var, width=12)
            self.end_date_entry.pack(side="left", padx=(5, 15))

        # Clear button
        ttk.Button(date_row, text="Clear", command=self._clear_dates).pack(side="left")

        # Help text
        ttk.Label(filter_frame, text="Leave dates empty to export all active customers",
                  foreground="gray").pack(anchor="w", pady=(10, 0))

        # Preview info
        self.preview_label = ttk.Label(main_frame, text="")
        self.preview_label.pack(anchor="w", pady=(0, 10))

        # Update preview on date change
        if CALENDAR_AVAILABLE:
            self.start_date_entry.bind("<<DateEntrySelected>>", lambda e: self._update_preview())
            self.end_date_entry.bind("<<DateEntrySelected>>", lambda e: self._update_preview())
        self.after(100, self._update_preview)

        # Buttons
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill="x", pady=(10, 0))

        ttk.Button(btn_frame, text="Export...", command=self._export).pack(side="left", padx=(0, 10))
        ttk.Button(btn_frame, text="Cancel", command=self.destroy).pack(side="left")

    def _clear_dates(self):
        """Clear date entries."""
        if CALENDAR_AVAILABLE:
            self.start_date_entry.delete(0, "end")
            self.end_date_entry.delete(0, "end")
        else:
            self.start_date_var.set("")
            self.end_date_var.set("")
        self._update_preview()

    def _get_date(self, entry) -> Optional[date]:
        """Get date from entry widget."""
        if CALENDAR_AVAILABLE:
            try:
                date_str = entry.get()
                if not date_str:
                    return None
                return entry.get_date()
            except:
                return None
        else:
            date_str = entry.get().strip()
            if not date_str:
                return None
            for fmt in ['%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d']:
                try:
                    return datetime.strptime(date_str, fmt).date()
                except ValueError:
                    continue
            return None

    def _update_preview(self):
        """Update the preview count."""
        start_date = self._get_date(self.start_date_entry)
        end_date = self._get_date(self.end_date_entry)

        contacts = db.get_customers_for_export(start_date, end_date)
        count = len(contacts)

        self.preview_label.config(text=f"Found {count} contact{'s' if count != 1 else ''} to export")

    def _export(self):
        """Perform the export."""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl is required for Excel export.\nInstall with: pip install openpyxl")
            return

        start_date = self._get_date(self.start_date_entry)
        end_date = self._get_date(self.end_date_entry)

        # Validate dates
        if start_date and end_date and start_date > end_date:
            messagebox.showerror("Error", "Start date cannot be after end date.")
            return

        # Get contacts
        contacts = db.get_customers_for_export(start_date, end_date)

        if not contacts:
            messagebox.showwarning("No Data", "No contacts found matching the criteria.")
            return

        # Choose output file
        default_name = f"Contact_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        output_path = filedialog.asksaveasfilename(
            title="Save Excel File",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
            initialfile=default_name,
            parent=self
        )

        if not output_path:
            return

        try:
            count = export_to_excel(contacts, output_path, start_date, end_date)

            messagebox.showinfo(
                "Export Complete",
                f"Successfully exported {count} contacts to:\n{output_path}",
                parent=self
            )

            # Offer to open the file
            if messagebox.askyesno("Open File?", "Would you like to open the exported file?", parent=self):
                os.startfile(output_path)

            self.destroy()

        except Exception as e:
            messagebox.showerror("Export Error", str(e), parent=self)
